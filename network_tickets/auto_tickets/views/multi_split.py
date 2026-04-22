from auto_tickets.forms_multisplit import IPDBFORM_MULTISPLIT
from auto_tickets.models import EomsTicketCreationTask
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST, require_GET
from django.views.decorators.csrf import csrf_exempt
from auto_tickets.tools import tickets_split
from auto_tickets.views.ITSR_Tools.eoms_automation_2 import create_ticket
from auto_tickets.views.ITSR_Tools.itsr_create import (
    create_ticket_session as itsr_create_ticket_session,
    submit_credentials as itsr_submit_credentials,
    submit_sms_code as itsr_submit_sms_code,
    wait_create_result as itsr_wait_create_result,
    cancel_session as itsr_cancel_session,
    get_session_status as itsr_get_session_status,
)
import openpyxl
import re
import asyncio
import json
import logging
import threading
import uuid
import time
import math
import shutil
import os
import glob as glob_module

logger = logging.getLogger(__name__)


def _eoms_task_row_to_dict(task):
    """Shape expected by multi_split.html poll (matches former in-memory task dict)."""
    if not task:
        return None
    return {
        'status': task.status,
        'success': task.success,
        'department': task.department,
        'requestor': task.requestor,
        'inst_id': task.inst_id or None,
        'message': task.message or '',
        'error': task.error or '',
        'need_captcha': task.need_captcha,
        'resume_token': task.cas_resume_token or None,
    }


def _session_get_ticket_title(django_session_key):
    """Read Excel-derived ticket title from Django session (works from background thread)."""
    if not django_session_key:
        return ''
    try:
        from django.contrib.sessions.backends.db import SessionStore

        store = SessionStore(session_key=django_session_key)
        store.load()
        return (store.get('last_ticket_title') or '').strip()
    except Exception:
        return ''


def _eoms_clear_session_cooldown(django_session_key, target_department):
    """Clear per-dept EOMS start cooldown (after success, failure, captcha, or exception)."""
    if not django_session_key or not target_department:
        return
    try:
        from django.contrib.sessions.backends.db import SessionStore

        store = SessionStore(session_key=django_session_key)
        store.load()
        k = f'ticket_creation_{target_department.lower()}_timestamp'
        if k in store:
            del store[k]
        store.save()
    except Exception as ex:
        logger.warning('EOMS: could not clear session cooldown: %s', ex)

# VPN source IP ranges (fixed for all VPN tickets).
# Both must exist in IPDB with the same location/device (e.g. SZ-VPN, DMZ SW01) or only the one in IPDB will appear in generated ticket files.
VPN_SOURCE_IPS = ['10.51.203.0/24', '10.51.204.0/24']


def _find_ticket_title_column_1based(sheet, default_col):
    """
    Locate the "Ticket Title" column (English or Chinese header in row 1 or 2).
    Falls back to default_col when not found (last column in updated samples).
    """
    for r in (1, 2):
        for c in range(1, 26):
            raw = sheet.cell(row=r, column=c).value
            if raw is None:
                continue
            v = str(raw).strip()
            vl = v.lower()
            if vl == 'ticket title' or 'ticket title' in vl:
                return c
            if '工单标题' in v:
                return c
    return default_col


def _detect_file_format(sheet):
    """
    Auto-detect whether the uploaded Excel file is ITSR format or VPN format.
    
    ITSR format: Has 'Source' in column B or C headers (e.g. 'Source Node Name', 'Source Node IP')
    VPN format: Column C header contains 'Description', Column D contains 'Protocol', Column E contains 'Port'
    
    Both formats have 'Ticket No.' in column A, so we differentiate by other columns.
    
    Returns: 'vpn' or 'itsr'
    """
    header_b = str(sheet.cell(row=1, column=2).value or '').strip().lower()
    header_c = str(sheet.cell(row=1, column=3).value or '').strip().lower()
    header_d = str(sheet.cell(row=1, column=4).value or '').strip().lower()
    header_e = str(sheet.cell(row=1, column=5).value or '').strip().lower()
    # Also check row 2 (Chinese headers) as fallback
    header_b_row2 = str(sheet.cell(row=2, column=2).value or '').strip().lower()
    header_c_row2 = str(sheet.cell(row=2, column=3).value or '').strip().lower()
    header_d_row2 = str(sheet.cell(row=2, column=4).value or '').strip().lower()
    header_e_row2 = str(sheet.cell(row=2, column=5).value or '').strip().lower()
    
    # ITSR: Column B = "Source Node Name", Column C = "Source Node IP"
    if 'source' in header_b or 'source' in header_c or '源' in header_b_row2 or '源' in header_c_row2:
        return 'itsr'
    
    # VPN: Column C = "Description", Column D = "Protocol"
    if 'description' in header_c or 'protocol' in header_d or '协议' in header_d or '协议' in header_d_row2:
        return 'vpn'
    
    # Fallback: check column E for port-related keywords
    if 'port' in header_e or '端口' in header_e_row2:
        return 'vpn'
    
    # Default to ITSR
    return 'itsr'


def _check_cloud_sn(ticket_list):
    """
    Check if a ticket list from tickets_split indicates Cloud or SN departments.
    Returns (needs_cloud, needs_sn) tuple.
    """
    needs_cloud = False
    needs_sn = False
    for ticket in ticket_list:
        ticket_lower = ticket.lower()
        if 'cloud' in ticket_lower:
            needs_cloud = True
        if 'sn' in ticket_lower:
            needs_sn = True
    return needs_cloud, needs_sn


def _check_itsr(ticket_list):
    """
    Check if a ticket list from tickets_split indicates an ITSR requirement.
    Returns True if any ticket contains 'itsr' (e.g., 'ITSR').
    """
    for ticket in ticket_list:
        if 'itsr' in ticket.lower():
            return True
    return False


# Directory for per-session EOMS files
EOMS_TEMPLATE_DIR = 'auto_tickets/views/EOMS_Ticket_file'
EOMS_SESSION_DIR = os.path.join(EOMS_TEMPLATE_DIR, 'session_files')

# Directory for per-session ITSR files
ITSR_TEMPLATE_DIR = 'auto_tickets/views/ITSR_Tools'
ITSR_SESSION_DIR = os.path.join(ITSR_TEMPLATE_DIR, 'itsr_session_files')


def _ensure_session_dir():
    """Ensure the per-session files directory exists."""
    os.makedirs(EOMS_SESSION_DIR, exist_ok=True)


def _write_eoms_template(template_name, source_name_dic, sip_dic, dest_name_dic, dip_dic, dport_dic, protocol_dic, requestor_dic, session_id=None):
    """
    Write data to a per-session copy of the EOMS template xlsx (Cloud or SN).
    
    Instead of modifying the shared template in-place, this copies the template
    to a unique per-session file and writes data there to avoid race conditions
    when multiple users process files simultaneously.
    
    Args:
        template_name: 'cloud' or 'sn' (used to pick the right template)
        source_name_dic, sip_dic, dest_name_dic, dip_dic, dport_dic, protocol_dic, requestor_dic: data dicts
        session_id: unique identifier for this session/request (uses uuid if None)
    
    Returns:
        str: path to the unique per-session file that was created
    """
    _ensure_session_dir()
    
    if session_id is None:
        session_id = str(uuid.uuid4())
    
    template_path = os.path.join(EOMS_TEMPLATE_DIR, f'eoms_{template_name}_template.xlsx')
    unique_filename = f'eoms_{template_name}_{session_id}.xlsx'
    unique_path = os.path.join(EOMS_SESSION_DIR, unique_filename)
    
    # Copy the clean template to a unique per-session file
    shutil.copy2(template_path, unique_path)
    
    # Write data to the unique copy
    wb = openpyxl.load_workbook(unique_path)
    sheet = wb.active
    # Clear existing data rows (from row 4 onwards) just in case
    for row in range(4, sheet.max_row + 1):
        for col in [2, 3, 4, 5, 6, 7, 8]:  # Columns B, C, D, E, F, G, H
            sheet.cell(row=row, column=col).value = None
    # Write new data
    for row_num in sip_dic.keys():
        sheet.cell(row=row_num, column=2).value = source_name_dic.get(row_num, '')
        sheet.cell(row=row_num, column=3).value = sip_dic[row_num]
        sheet.cell(row=row_num, column=4).value = dest_name_dic.get(row_num, '')
        sheet.cell(row=row_num, column=5).value = dip_dic.get(row_num, '')
        dport_val = dport_dic.get(row_num, '')
        protocol_val = protocol_dic.get(row_num, '')
        sheet.cell(row=row_num, column=6).value = '\n '.join(dport_val) if isinstance(dport_val, list) else dport_val
        sheet.cell(row=row_num, column=7).value = '\n '.join(protocol_val) if isinstance(protocol_val, list) else protocol_val
        sheet.cell(row=row_num, column=8).value = requestor_dic.get(row_num, '')
    wb.save(unique_path)
    
    return unique_path


def _cleanup_old_session_files(max_age_seconds=3600):
    """
    Remove per-session EOMS files older than max_age_seconds (default: 1 hour).
    Called opportunistically during file processing to prevent accumulation.
    """
    try:
        if not os.path.exists(EOMS_SESSION_DIR):
            return
        now = time.time()
        for filepath in glob_module.glob(os.path.join(EOMS_SESSION_DIR, 'eoms_*.xlsx')):
            try:
                file_age = now - os.path.getmtime(filepath)
                if file_age > max_age_seconds:
                    os.remove(filepath)
            except OSError:
                pass  # File may have been already deleted by another process
    except Exception:
        pass  # Cleanup is best-effort, never block the main flow


def _cleanup_session_file(file_path):
    """Remove a specific session file after it's been used."""
    try:
        if file_path and os.path.exists(file_path) and EOMS_SESSION_DIR in file_path:
            os.remove(file_path)
    except OSError:
        pass  # Best-effort cleanup


# ============================================================
# ITSR per-session file helpers
# ============================================================

def _ensure_itsr_session_dir():
    """Ensure the per-session ITSR files directory exists."""
    os.makedirs(ITSR_SESSION_DIR, exist_ok=True)


def _write_itsr_template(source_name_dic, sip_dic, dest_name_dic, dip_dic, dport_dic, protocol_dic, requestor_dic, session_id=None):
    """
    Create a per-session ITSR attachment xlsx with the collected ITSR data.
    
    The file is created from scratch (no pre-existing template needed).
    
    Args:
        source_name_dic, sip_dic, dest_name_dic, dip_dic, dport_dic, protocol_dic, requestor_dic: data dicts
        session_id: unique identifier for this session/request (uses uuid if None)
    
    Returns:
        str: absolute path to the unique per-session file that was created
    """
    _ensure_itsr_session_dir()

    if session_id is None:
        session_id = str(uuid.uuid4())

    unique_filename = f'itsr_{session_id}.xlsx'
    unique_path = os.path.join(ITSR_SESSION_DIR, unique_filename)

    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "ITSR Network Policy"

    # Row 1: English headers
    headers = [
        'No.', 'Source Node Name', 'Source Node IP',
        'Destination Node Name', 'Destination Node IP',
        'Destination Port', 'Protocol', 'Staff Number',
    ]
    for col, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col).value = header

    # Row 2: Chinese headers
    chinese_headers = [
        '序号', '源节点名称', '源节点IP',
        '目标节点名称', '目标节点IP',
        '目标端口', '协议', '申请人工号',
    ]
    for col, header in enumerate(chinese_headers, 1):
        sheet.cell(row=2, column=col).value = header

    # Row 3: reserved (example row placeholder)
    sheet.cell(row=3, column=1).value = 'Example'

    # Data rows starting from row 4 (matching EOMS convention)
    for row_num in sorted(sip_dic.keys()):
        idx = row_num - 3  # 1-based index
        sheet.cell(row=row_num, column=1).value = idx
        sheet.cell(row=row_num, column=2).value = source_name_dic.get(row_num, '')
        sheet.cell(row=row_num, column=3).value = sip_dic[row_num]
        sheet.cell(row=row_num, column=4).value = dest_name_dic.get(row_num, '')
        sheet.cell(row=row_num, column=5).value = dip_dic.get(row_num, '')
        dport_val = dport_dic.get(row_num, '')
        protocol_val = protocol_dic.get(row_num, '')
        sheet.cell(row=row_num, column=6).value = (
            '\n '.join(dport_val) if isinstance(dport_val, list) else dport_val
        )
        sheet.cell(row=row_num, column=7).value = (
            '\n '.join(protocol_val) if isinstance(protocol_val, list) else protocol_val
        )
        sheet.cell(row=row_num, column=8).value = requestor_dic.get(row_num, '')

    wb.save(unique_path)
    # Return the absolute path so itsr_create.py can find it
    return os.path.abspath(unique_path)


def _cleanup_old_itsr_session_files(max_age_seconds=3600):
    """
    Remove per-session ITSR files older than max_age_seconds (default: 1 hour).
    Called opportunistically during file processing to prevent accumulation.
    """
    try:
        if not os.path.exists(ITSR_SESSION_DIR):
            return
        now = time.time()
        for pattern in ('itsr_*.xlsx', 'original_*'):
            for filepath in glob_module.glob(os.path.join(ITSR_SESSION_DIR, pattern)):
                try:
                    file_age = now - os.path.getmtime(filepath)
                    if file_age > max_age_seconds:
                        os.remove(filepath)
                except OSError:
                    pass
    except Exception:
        pass  # Cleanup is best-effort


def _cleanup_itsr_session_file(file_path):
    """Remove a specific ITSR session file after it's been used."""
    try:
        if file_path and os.path.exists(file_path) and ITSR_SESSION_DIR in file_path:
            os.remove(file_path)
    except OSError:
        pass  # Best-effort cleanup


def _sanitize_upload_basename(name):
    """Safe filename fragment for storing a copy of the user's upload under itsr_session_files."""
    base = os.path.basename(name or '') or 'upload.xlsx'
    base = re.sub(r'[^\w.\-()\u4e00-\u9fff]', '_', base, flags=re.UNICODE)
    return base[:180] if len(base) > 180 else base


def _save_itsr_original_upload(uploaded_file, processing_session_id):
    """
    Save the user's workbook next to the generated ITSR xlsx so BPM can attach both.
    Call after parsing; uses seek(0) because openpyxl may have read the upload stream.
    """
    _ensure_itsr_session_dir()
    safe = _sanitize_upload_basename(uploaded_file.name)
    if not any(safe.lower().endswith(ext) for ext in ('.xlsx', '.xls', '.xlsm')):
        root, ext = os.path.splitext(safe)
        safe = (root or 'upload') + (ext if ext else '.xlsx')
    unique_filename = f'original_{processing_session_id}_{safe}'
    unique_path = os.path.join(ITSR_SESSION_DIR, unique_filename)
    try:
        uploaded_file.seek(0)
        with open(unique_path, 'wb') as out:
            for chunk in uploaded_file.chunks():
                out.write(chunk)
        return os.path.abspath(unique_path)
    except OSError:
        logger.warning('Could not save ITSR original upload copy', exc_info=True)
        return None


def _process_itsr_file(sheet):
    """
    Process an ITSR format Excel file.
    Returns dict with: result_list, detected_cloud, detected_sn, detected_itsr, cloud/sn/itsr dicts, staff_number.
    """
    pattern = r'[ ,\n、]'
    result_list = []

    sn_source_name_dic = {}
    sn_sip_dic = {}
    sn_dest_name_dic = {}
    sn_dip_dic = {}
    sn_dport_dic = {}
    sn_protocol_dic = {}
    sn_requestor_dic = {}

    cloud_source_name_dic = {}
    cloud_sip_dic = {}
    cloud_dest_name_dic = {}
    cloud_dip_dic = {}
    cloud_dport_dic = {}
    cloud_protocol_dic = {}
    cloud_requestor_dic = {}

    itsr_source_name_dic = {}
    itsr_sip_dic = {}
    itsr_dest_name_dic = {}
    itsr_dip_dic = {}
    itsr_dport_dic = {}
    itsr_protocol_dic = {}
    itsr_requestor_dic = {}

    detected_cloud = False
    detected_sn = False
    detected_itsr = False
    cloud_num = 4
    sn_num = 4
    itsr_num = 4
    cloud_processed_pairs = set()
    sn_processed_pairs = set()
    itsr_processed_pairs = set()
    staff_number = ''
    ticket_title = ''
    title_col = _find_ticket_title_column_1based(sheet, 9)

    for row_num, row in enumerate(sheet.iter_rows(min_row=4, max_col=9), start=4):
        try:
            if row[2].value and row[4].value and row[5].value and row[6].value:
                source_name = str(row[1].value).strip() if row[1].value else ''
                source_ip_list = [item for item in re.split(pattern, str(row[2].value)) if item.strip()]
                dest_name = str(row[3].value).strip() if row[3].value else ''
                destination_ip_list = [item for item in re.split(pattern, str(row[4].value)) if item.strip()]
                destination_port_list = [item for item in re.split(pattern, str(row[5].value)) if item.strip()]
                protocol_list = [item for item in re.split(pattern, str(row[6].value)) if item.strip()]
                requestor_list = [item for item in re.split(pattern, str(row[7].value if row[7].value else '')) if item.strip()]
                if requestor_list and not staff_number:
                    staff_number = requestor_list[0]
                tv = sheet.cell(row=row_num, column=title_col).value
                if tv is not None and str(tv).strip() and not ticket_title:
                    ticket_title = str(tv).strip()
                for i in source_ip_list:
                    source_ip = i.strip().replace('\u200b', '')
                    judge_source_ip = re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', source_ip)
                    if judge_source_ip:
                        for destination_ip in destination_ip_list:
                            destination_ip = destination_ip.strip().replace('\u200b', '')
                            judge_destination_ip = re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', destination_ip)
                            if judge_destination_ip:
                                try:
                                    ticket_list = tickets_split(source_ip, destination_ip, return_list=True)
                                    needs_cloud, needs_sn = _check_cloud_sn(ticket_list)
                                    needs_itsr = _check_itsr(ticket_list)
                                    if needs_cloud:
                                        detected_cloud = True
                                    if needs_sn:
                                        detected_sn = True
                                    if needs_itsr:
                                        detected_itsr = True

                                    sorted_ports = sorted(destination_port_list)
                                    sorted_protocols = sorted(protocol_list)
                                    unique_key = f"{source_ip}_{destination_ip}_{','.join(sorted_protocols)}_{','.join(sorted_ports)}"

                                    if needs_cloud and unique_key not in cloud_processed_pairs:
                                        cloud_source_name_dic[cloud_num] = source_name
                                        cloud_sip_dic[cloud_num] = source_ip
                                        cloud_dest_name_dic[cloud_num] = dest_name
                                        cloud_dip_dic[cloud_num] = destination_ip
                                        cloud_dport_dic[cloud_num] = destination_port_list
                                        cloud_protocol_dic[cloud_num] = protocol_list
                                        cloud_requestor_dic[cloud_num] = requestor_list[0] if requestor_list else ''
                                        cloud_processed_pairs.add(unique_key)
                                        cloud_num += 1

                                    if needs_sn and unique_key not in sn_processed_pairs:
                                        sn_source_name_dic[sn_num] = source_name
                                        sn_sip_dic[sn_num] = source_ip
                                        sn_dest_name_dic[sn_num] = dest_name
                                        sn_dip_dic[sn_num] = destination_ip
                                        sn_dport_dic[sn_num] = destination_port_list
                                        sn_protocol_dic[sn_num] = protocol_list
                                        sn_requestor_dic[sn_num] = requestor_list[0] if requestor_list else ''
                                        sn_processed_pairs.add(unique_key)
                                        sn_num += 1

                                    if needs_itsr and unique_key not in itsr_processed_pairs:
                                        itsr_source_name_dic[itsr_num] = source_name
                                        itsr_sip_dic[itsr_num] = source_ip
                                        itsr_dest_name_dic[itsr_num] = dest_name
                                        itsr_dip_dic[itsr_num] = destination_ip
                                        itsr_dport_dic[itsr_num] = destination_port_list
                                        itsr_protocol_dic[itsr_num] = protocol_list
                                        itsr_requestor_dic[itsr_num] = requestor_list[0] if requestor_list else ''
                                        itsr_processed_pairs.add(unique_key)
                                        itsr_num += 1

                                    result_str = tickets_split(source_ip, destination_ip, return_list=False)
                                    result_list.append(result_str)

                                except Exception as e:
                                    error_msg = f"ERROR: Row {row_num} - Failed to process {source_ip} to {destination_ip}: {str(e)}"
                                    result_list.append(error_msg)
                            else:
                                error_msg = f"ERROR: Row {row_num} - Invalid destination IP format: {destination_ip}"
                                result_list.append(error_msg)
                    else:
                        error_msg = f"ERROR: Row {row_num} - Invalid source IP format: {source_ip}"
                        result_list.append(error_msg)
        except Exception as e:
            error_msg = f"ERROR: Row {row_num} - Error processing row: {str(e)}"
            result_list.append(error_msg)

    return {
        'result_list': result_list,
        'detected_cloud': detected_cloud,
        'detected_sn': detected_sn,
        'detected_itsr': detected_itsr,
        'sn_source_name_dic': sn_source_name_dic, 'sn_sip_dic': sn_sip_dic,
        'sn_dest_name_dic': sn_dest_name_dic, 'sn_dip_dic': sn_dip_dic,
        'sn_dport_dic': sn_dport_dic, 'sn_protocol_dic': sn_protocol_dic,
        'sn_requestor_dic': sn_requestor_dic,
        'cloud_source_name_dic': cloud_source_name_dic, 'cloud_sip_dic': cloud_sip_dic,
        'cloud_dest_name_dic': cloud_dest_name_dic, 'cloud_dip_dic': cloud_dip_dic,
        'cloud_dport_dic': cloud_dport_dic, 'cloud_protocol_dic': cloud_protocol_dic,
        'cloud_requestor_dic': cloud_requestor_dic,
        'itsr_source_name_dic': itsr_source_name_dic, 'itsr_sip_dic': itsr_sip_dic,
        'itsr_dest_name_dic': itsr_dest_name_dic, 'itsr_dip_dic': itsr_dip_dic,
        'itsr_dport_dic': itsr_dport_dic, 'itsr_protocol_dic': itsr_protocol_dic,
        'itsr_requestor_dic': itsr_requestor_dic,
        'staff_number': staff_number,
        'ticket_title': ticket_title,
    }


def _process_vpn_file(sheet):
    """
    Process a VPN Network Ticket format Excel file.
    
    VPN format columns (data starts at row 4, row 3 is example):
        A (1): Ticket No.
        B (2): Destination IP
        C (3): Description (optional)
        D (4): Protocol (TCP/UDP/ICMP)
        E (5): Port/Services No.
        F (6): Vendor Name
        G (7): Requested by / Staff Number
    
    Source IPs are fixed: 10.51.203.0/24, 10.51.204.0/24
    
    Returns dict with: result_list, detected_cloud, detected_sn, detected_itsr, cloud/sn/itsr dicts, staff_number.
    """
    pattern = r'[ ,\n、，]'
    result_list = []

    sn_source_name_dic = {}
    sn_sip_dic = {}
    sn_dest_name_dic = {}
    sn_dip_dic = {}
    sn_dport_dic = {}
    sn_protocol_dic = {}
    sn_requestor_dic = {}

    cloud_source_name_dic = {}
    cloud_sip_dic = {}
    cloud_dest_name_dic = {}
    cloud_dip_dic = {}
    cloud_dport_dic = {}
    cloud_protocol_dic = {}
    cloud_requestor_dic = {}

    itsr_source_name_dic = {}
    itsr_sip_dic = {}
    itsr_dest_name_dic = {}
    itsr_dip_dic = {}
    itsr_dport_dic = {}
    itsr_protocol_dic = {}
    itsr_requestor_dic = {}

    detected_cloud = False
    detected_sn = False
    detected_itsr = False
    cloud_num = 4
    sn_num = 4
    itsr_num = 4
    cloud_processed_pairs = set()
    sn_processed_pairs = set()
    itsr_processed_pairs = set()
    staff_number = ''
    ticket_title = ''
    title_col = _find_ticket_title_column_1based(sheet, 8)

    for row_num, row in enumerate(sheet.iter_rows(min_row=4, max_col=8), start=4):
        try:
            # VPN format: B=Dest IP, D=Protocol, E=Port are required (C=Description is optional)
            if not (row[1].value and row[3].value and row[4].value):
                continue

            destination_ip_list = [item for item in re.split(pattern, str(row[1].value)) if item.strip()]
            dest_name = str(row[2].value).strip() if row[2].value else ''
            protocol_list = [item for item in re.split(pattern, str(row[3].value)) if item.strip()]
            destination_port_list = [item for item in re.split(pattern, str(row[4].value)) if item.strip()]
            # Column G (index 6): Staff Number (requestor/originator)
            row_staff = str(row[6].value).strip() if row[6].value else ''
            if row_staff and not staff_number:
                staff_number = row_staff
            tv = sheet.cell(row=row_num, column=title_col).value
            if tv is not None and str(tv).strip() and not ticket_title:
                ticket_title = str(tv).strip()

            for destination_ip_raw in destination_ip_list:
                destination_ip = destination_ip_raw.strip().replace('\u200b', '')
                judge_destination_ip = re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', destination_ip)
                if not judge_destination_ip:
                    error_msg = f"ERROR: Row {row_num} - Invalid destination IP format: {destination_ip}"
                    result_list.append(error_msg)
                    continue

                for source_ip in VPN_SOURCE_IPS:
                    try:
                        ticket_list = tickets_split(source_ip, destination_ip, return_list=True)
                        needs_cloud, needs_sn = _check_cloud_sn(ticket_list)
                        needs_itsr = _check_itsr(ticket_list)
                        if needs_cloud:
                            detected_cloud = True
                        if needs_sn:
                            detected_sn = True
                        if needs_itsr:
                            detected_itsr = True

                        sorted_ports = sorted(destination_port_list)
                        sorted_protocols = sorted(protocol_list)
                        unique_key = f"{source_ip}_{destination_ip}_{','.join(sorted_protocols)}_{','.join(sorted_ports)}"

                        if needs_cloud and unique_key not in cloud_processed_pairs:
                            cloud_source_name_dic[cloud_num] = 'VPN System'
                            cloud_sip_dic[cloud_num] = source_ip
                            cloud_dest_name_dic[cloud_num] = dest_name
                            cloud_dip_dic[cloud_num] = destination_ip
                            cloud_dport_dic[cloud_num] = destination_port_list
                            cloud_protocol_dic[cloud_num] = protocol_list
                            cloud_requestor_dic[cloud_num] = staff_number
                            cloud_processed_pairs.add(unique_key)
                            cloud_num += 1

                        if needs_sn and unique_key not in sn_processed_pairs:
                            sn_source_name_dic[sn_num] = 'VPN System'
                            sn_sip_dic[sn_num] = source_ip
                            sn_dest_name_dic[sn_num] = dest_name
                            sn_dip_dic[sn_num] = destination_ip
                            sn_dport_dic[sn_num] = destination_port_list
                            sn_protocol_dic[sn_num] = protocol_list
                            sn_requestor_dic[sn_num] = staff_number
                            sn_processed_pairs.add(unique_key)
                            sn_num += 1

                        if needs_itsr and unique_key not in itsr_processed_pairs:
                            itsr_source_name_dic[itsr_num] = 'VPN System'
                            itsr_sip_dic[itsr_num] = source_ip
                            itsr_dest_name_dic[itsr_num] = dest_name
                            itsr_dip_dic[itsr_num] = destination_ip
                            itsr_dport_dic[itsr_num] = destination_port_list
                            itsr_protocol_dic[itsr_num] = protocol_list
                            itsr_requestor_dic[itsr_num] = staff_number
                            itsr_processed_pairs.add(unique_key)
                            itsr_num += 1

                        result_str = tickets_split(source_ip, destination_ip, return_list=False)
                        result_list.append(result_str)

                    except Exception as e:
                        error_msg = f"ERROR: Row {row_num} - Failed to process {source_ip} to {destination_ip}: {str(e)}"
                        result_list.append(error_msg)

        except Exception as e:
            error_msg = f"ERROR: Row {row_num} - Error processing row: {str(e)}"
            result_list.append(error_msg)

    return {
        'result_list': result_list,
        'detected_cloud': detected_cloud,
        'detected_sn': detected_sn,
        'detected_itsr': detected_itsr,
        'sn_source_name_dic': sn_source_name_dic, 'sn_sip_dic': sn_sip_dic,
        'sn_dest_name_dic': sn_dest_name_dic, 'sn_dip_dic': sn_dip_dic,
        'sn_dport_dic': sn_dport_dic, 'sn_protocol_dic': sn_protocol_dic,
        'sn_requestor_dic': sn_requestor_dic,
        'cloud_source_name_dic': cloud_source_name_dic, 'cloud_sip_dic': cloud_sip_dic,
        'cloud_dest_name_dic': cloud_dest_name_dic, 'cloud_dip_dic': cloud_dip_dic,
        'cloud_dport_dic': cloud_dport_dic, 'cloud_protocol_dic': cloud_protocol_dic,
        'cloud_requestor_dic': cloud_requestor_dic,
        'itsr_source_name_dic': itsr_source_name_dic, 'itsr_sip_dic': itsr_sip_dic,
        'itsr_dest_name_dic': itsr_dest_name_dic, 'itsr_dip_dic': itsr_dip_dic,
        'itsr_dport_dic': itsr_dport_dic, 'itsr_protocol_dic': itsr_protocol_dic,
        'itsr_requestor_dic': itsr_requestor_dic,
        'staff_number': staff_number,
        'ticket_title': ticket_title,
    }


def multi_split(request):
    if request.method == 'POST':
        # File processing logic (supports both ITSR and VPN formats)
        form = IPDBFORM_MULTISPLIT(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['file']
            try:
                wb = openpyxl.load_workbook(uploaded_file)
                sheet = wb.active

                # Auto-detect file format
                file_format = _detect_file_format(sheet)
                print(f"[DEBUG] Detected file format: {file_format}")

                # Process based on detected format
                if file_format == 'vpn':
                    data = _process_vpn_file(sheet)
                else:
                    data = _process_itsr_file(sheet)

                result_list = data['result_list']
                detected_cloud = data['detected_cloud']
                detected_sn = data['detected_sn']
                detected_itsr = data.get('detected_itsr', False)
                staff_number = data.get('staff_number', '')
                ticket_title = (data.get('ticket_title') or '').strip()

                # Opportunistically clean up old session files
                _cleanup_old_session_files()
                _cleanup_old_itsr_session_files()

                # Generate a unique session ID for this file processing request
                processing_session_id = str(uuid.uuid4())

                # Clean up previous session files if they exist
                old_sn_path = request.session.get('eoms_sn_file_path')
                old_cloud_path = request.session.get('eoms_cloud_file_path')
                old_itsr_path = request.session.get('itsr_file_path')
                old_itsr_original = request.session.get('itsr_original_file_path')
                if old_sn_path:
                    _cleanup_session_file(old_sn_path)
                if old_cloud_path:
                    _cleanup_session_file(old_cloud_path)
                if old_itsr_path:
                    _cleanup_itsr_session_file(old_itsr_path)
                if old_itsr_original:
                    _cleanup_itsr_session_file(old_itsr_original)

                # Write EOMS SN template (per-session unique file)
                eoms_sn_path = None
                if data['sn_sip_dic']:
                    eoms_sn_path = _write_eoms_template(
                        'sn',
                        data['sn_source_name_dic'], data['sn_sip_dic'],
                        data['sn_dest_name_dic'], data['sn_dip_dic'],
                        data['sn_dport_dic'], data['sn_protocol_dic'],
                        data['sn_requestor_dic'],
                        session_id=processing_session_id
                    )

                # Write EOMS Cloud template (per-session unique file)
                eoms_cloud_path = None
                if data['cloud_sip_dic']:
                    eoms_cloud_path = _write_eoms_template(
                        'cloud',
                        data['cloud_source_name_dic'], data['cloud_sip_dic'],
                        data['cloud_dest_name_dic'], data['cloud_dip_dic'],
                        data['cloud_dport_dic'], data['cloud_protocol_dic'],
                        data['cloud_requestor_dic'],
                        session_id=processing_session_id
                    )

                # Write ITSR template (per-session unique file)
                itsr_path = None
                itsr_original_path = None
                if data['itsr_sip_dic']:
                    itsr_path = _write_itsr_template(
                        data['itsr_source_name_dic'], data['itsr_sip_dic'],
                        data['itsr_dest_name_dic'], data['itsr_dip_dic'],
                        data['itsr_dport_dic'], data['itsr_protocol_dic'],
                        data['itsr_requestor_dic'],
                        session_id=processing_session_id
                    )
                    # Also persist the workbook the user uploaded (e.g. VPN template) as a second BPM attachment
                    itsr_original_path = _save_itsr_original_upload(
                        uploaded_file, processing_session_id
                    )

                # Check if we got any results
                if result_list:
                    # Process the result list to separate errors from other messages
                    error_messages = []
                    
                    for message in result_list:
                        # More specific error detection to avoid false positives
                        if message and any(keyword in message.lower() for keyword in ['failed:', 'traceback:', 'validation failed', 'connection failed', 'error:']):
                            error_messages.append(message)
                    
                    # Store results in session for ticket creation redirect
                    request.session['last_result_list'] = result_list
                    request.session['last_error_messages'] = error_messages
                    request.session['last_has_errors'] = len(error_messages) > 0
                    request.session['last_show_cloud_button'] = detected_cloud
                    request.session['last_show_sn_button'] = detected_sn
                    request.session['last_show_itsr_button'] = detected_itsr
                    # Store first requestor for each department
                    request.session['last_cloud_requestor'] = data['cloud_requestor_dic'].get(4, '') if data['cloud_requestor_dic'] else ''
                    request.session['last_sn_requestor'] = data['sn_requestor_dic'].get(4, '') if data['sn_requestor_dic'] else ''
                    # Store staff number (from VPN files)
                    request.session['last_staff_number'] = staff_number
                    request.session['last_ticket_title'] = ticket_title
                    # Store per-session unique file paths for ticket creation
                    request.session['eoms_cloud_file_path'] = eoms_cloud_path
                    request.session['eoms_sn_file_path'] = eoms_sn_path
                    request.session['itsr_file_path'] = itsr_path
                    if itsr_original_path:
                        request.session['itsr_original_file_path'] = itsr_original_path
                    else:
                        request.session.pop('itsr_original_file_path', None)
                    # Clear previous ticket creation state when processing new file
                    request.session.pop('ticket_cloud_created', None)
                    request.session.pop('ticket_sn_created', None)
                    request.session.pop('ticket_itsr_created', None)
                    request.session.pop('itsr_create_session_id', None)
                    request.session.save()
                    
                    # Return results with error messages and department detection
                    return render(request, 'multi_split.html', {
                        'result_list': result_list,
                        'error_messages': error_messages,
                        'has_errors': len(error_messages) > 0,
                        'show_cloud_button': detected_cloud,
                        'show_sn_button': detected_sn,
                        'show_itsr_button': detected_itsr,
                        'cloud_ticket_created': False,
                        'sn_ticket_created': False,
                        'itsr_ticket_created': False,
                        'staff_number': staff_number,
                        'file_format': file_format,
                    })
                else:
                    # No results found — form.file.errors is rendered in the upload card (see multi_split.html)
                    if file_format == 'vpn':
                        form.add_error(
                            'file',
                            'No valid data found in the VPN Excel file. Please check that your file has '
                            'Destination IP in column B, Description in column C, Protocol in column D, '
                            'Port in column E, Staff Number in column G, and optional Ticket Title in column H '
                            'starting from row 4.',
                        )
                    else:
                        form.add_error(
                            'file',
                            'No valid data found in the Excel file. Please check that your file has data in '
                            'columns C–G (source IP, destination IP, port, protocol, staff) starting from row 4, '
                            'and optional Ticket Title in the last column.',
                        )
                    return render(request, 'multi_split.html', {'form': form})
            
            except Exception as e:
                # If there's an error processing the file, show the form with error
                error_msg = f'Error processing file: {str(e)}'
                form.add_error('file', error_msg)
                return render(request, 'multi_split.html', {
                    'form': form,
                    'error_messages': [error_msg],
                    'has_errors': True
                })
        else:
            # Form is not valid, render with errors
            return render(request, 'multi_split.html', {'form': form})
    else:
        # GET request - always show empty form for fresh uploads
        # Results will be shown after POST file processing
        form = IPDBFORM_MULTISPLIT()
        return render(request, 'multi_split.html', {'form': form})


# ============================================================
# AJAX API Endpoints for ticket creation with background processing
# ============================================================

def _run_ticket_creation(
    task_id,
    target_department,
    file_path,
    requestor,
    django_session_key,
    username,
    password,
    captcha_code,
    resume_token,
):
    """Background worker: Playwright EOMS create (runs in a daemon thread)."""
    from django.db import close_old_connections

    close_old_connections()
    try:
        excel_title = _session_get_ticket_title(django_session_key)
        result = asyncio.run(
            create_ticket(
                target_department=target_department,
                file_path=file_path,
                username=username,
                password=password,
                originator=requestor if requestor else None,
                captcha_code=captcha_code or '',
                resume_token=resume_token or '',
                title=excel_title or None,
            )
        )

        if result.get('need_captcha'):
            EomsTicketCreationTask.objects.filter(task_id=task_id).update(
                status='need_captcha',
                success=False,
                need_captcha=True,
                error=result.get('error') or result.get('message') or 'Captcha required',
                cas_resume_token=result.get('resume_token') or '',
            )
            _eoms_clear_session_cooldown(django_session_key, target_department)
            return

        if result.get('success'):
            EomsTicketCreationTask.objects.filter(task_id=task_id).update(
                status='completed',
                success=True,
                need_captcha=False,
                inst_id=str(result.get('inst_id') or ''),
                message=result.get('message') or '',
                error='',
            )
            _cleanup_session_file(file_path)
            # Same as error/captcha paths: release cooldown so another EOMS (e.g. after ITSR) is not blocked
            _eoms_clear_session_cooldown(django_session_key, target_department)
        else:
            EomsTicketCreationTask.objects.filter(task_id=task_id).update(
                status='error',
                success=False,
                need_captcha=False,
                error=result.get('error') or result.get('message') or 'Ticket creation failed',
            )
            _eoms_clear_session_cooldown(django_session_key, target_department)
    except Exception as e:
        logger.exception('EOMS background task %s failed', task_id)
        EomsTicketCreationTask.objects.filter(task_id=task_id).update(
            status='error',
            success=False,
            need_captcha=False,
            error=str(e),
        )
        _eoms_clear_session_cooldown(django_session_key, target_department)
    finally:
        close_old_connections()


def _parse_eoms_create_payload(request):
    """Support JSON (modal) and form POST (legacy)."""
    ct = (request.content_type or '').lower()
    if 'application/json' in ct and request.body:
        try:
            return json.loads(request.body)
        except json.JSONDecodeError:
            return {}
    return request.POST


@require_POST
def api_create_eoms_ticket(request):
    """
    Start EOMS ticket creation in a background thread; returns task_id for polling.

    JSON body (preferred):
        username, password, target_department ("Cloud" | "SN"),
        captcha_code (optional), resume_token (optional, SMS second step)

    Form POST (legacy): same field names.
    """
    try:
        return _api_create_eoms_ticket_impl(request)
    except Exception as e:
        logger.exception('api_create_eoms_ticket failed')
        return JsonResponse(
            {
                'success': False,
                'error': 'An unexpected server error occurred. Refresh the page and try again, or check server logs.',
            },
            status=500,
        )


def _api_create_eoms_ticket_impl(request):
    data = _parse_eoms_create_payload(request)
    target_department = (data.get('target_department') or '').strip()
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    captcha_code = (data.get('captcha_code') or '').strip()
    resume_token = (data.get('resume_token') or '').strip()

    if target_department not in ['Cloud', 'SN']:
        return JsonResponse({'success': False, 'error': 'Invalid department'}, status=400)

    if not username or not password:
        return JsonResponse(
            {'success': False, 'error': 'Username and password are required.'},
            status=400,
        )

    cooldown_key = f'ticket_creation_{target_department.lower()}_timestamp'
    last_creation_time = request.session.get(cooldown_key, 0)
    current_time = time.time()
    cooldown_seconds = 30

    elapsed = current_time - last_creation_time
    if elapsed < cooldown_seconds:
        # ceil avoids "wait 0 seconds" when e.g. 29.9s elapsed (int() would truncate to 0)
        remaining_time = max(1, int(math.ceil(cooldown_seconds - elapsed)))
        return JsonResponse({
            'success': False,
            'error': f'Please wait {remaining_time} seconds before creating another ticket.',
            'cooldown': True,
            'remaining': remaining_time,
        })

    request.session[cooldown_key] = current_time
    request.session.save()

    if target_department == 'Cloud':
        file_path = request.session.get('eoms_cloud_file_path')
        requestor = request.session.get('last_cloud_requestor', '')
    else:
        file_path = request.session.get('eoms_sn_file_path')
        requestor = request.session.get('last_sn_requestor', '')

    if not file_path or not os.path.exists(file_path):
        request.session.pop(cooldown_key, None)
        request.session.save()
        return JsonResponse({
            'success': False,
            'error': 'Session expired or file not found. Please re-upload and process the file.',
        }, status=400)

    task_id = str(uuid.uuid4())
    EomsTicketCreationTask.objects.create(
        task_id=task_id,
        status='processing',
        department=target_department,
        requestor=requestor or '',
    )

    django_session_key = getattr(request.session, 'session_key', None) or ''

    thread = threading.Thread(
        target=_run_ticket_creation,
        args=(
            task_id,
            target_department,
            file_path,
            requestor,
            django_session_key,
            username,
            password,
            captcha_code,
            resume_token,
        ),
    )
    thread.daemon = True
    thread.start()

    return JsonResponse({
        'success': True,
        'task_id': task_id,
        'message': 'Ticket creation started',
    })


@require_GET
def api_check_ticket_status(request, task_id):
    """
    Poll background EOMS task status (DB-backed — safe across Gunicorn workers).
    """
    try:
        try:
            row = EomsTicketCreationTask.objects.get(task_id=task_id)
        except EomsTicketCreationTask.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Task not found'}, status=404)

        task = _eoms_task_row_to_dict(row)

        if row.status == 'completed' and row.success:
            dept_lower = row.department.lower()
            request.session[f'ticket_{dept_lower}_created'] = True
            file_key = f'eoms_{dept_lower}_file_path'
            request.session.pop(file_key, None)
            request.session.modified = True
            request.session.save()

        return JsonResponse({
            'success': True,
            'task': task,
        })
    except Exception as e:
        logger.exception('api_check_ticket_status failed for task_id=%s', task_id)
        return JsonResponse(
            {
                'success': False,
                'error': 'Could not read task status. Refresh the page and try again.',
            },
            status=500,
        )


@csrf_exempt
@require_POST
def api_create_eoms_ticket_v2(request):
    """
    EOMS ticket creation with user-provided credentials (blocking).
    Follows the same modal flow as ITSR: credentials → optional captcha → result.

    Expects JSON body:
        {
            "username": "...",
            "password": "...",
            "target_department": "Cloud" | "SN",
            "captcha_code": ""  (optional, for retry after captcha detected)
        }
    """
    try:
        data = json.loads(request.body) if request.body else {}
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        target_department = data.get('target_department', '').strip()
        captcha_code = data.get('captcha_code', '').strip()
        resume_token = data.get('resume_token', '').strip()

        if not username or not password:
            return JsonResponse({
                'success': False,
                'error': 'Username and password are required.',
            }, status=400)

        if target_department not in ['Cloud', 'SN']:
            return JsonResponse({
                'success': False,
                'error': 'Invalid department. Must be "Cloud" or "SN".',
            }, status=400)

        # Determine file path and requestor from session
        if target_department == 'Cloud':
            file_path = request.session.get('eoms_cloud_file_path')
            requestor = request.session.get('last_cloud_requestor', '')
        else:
            file_path = request.session.get('eoms_sn_file_path')
            requestor = request.session.get('last_sn_requestor', '')

        if not file_path or not os.path.exists(file_path):
            return JsonResponse({
                'success': False,
                'error': 'Session expired or file not found. Please re-upload and process the file.',
            }, status=400)

        excel_title = (request.session.get('last_ticket_title') or '').strip()
        result = asyncio.run(create_ticket(
            target_department=target_department,
            file_path=file_path,
            username=username,
            password=password,
            originator=requestor if requestor else None,
            captcha_code=captcha_code,
            resume_token=resume_token,
            title=excel_title or None,
        ))

        if result.get('need_captcha'):
            return JsonResponse({
                'success': False,
                'need_captcha': True,
                'error': result.get('error', 'Captcha/SMS verification required.'),
                'resume_token': result.get('resume_token'),
            })

        if result.get('success'):
            dept_lower = target_department.lower()
            request.session[f'ticket_{dept_lower}_created'] = True
            # Clean up per-session file
            _cleanup_session_file(file_path)
            file_key = f'eoms_{dept_lower}_file_path'
            request.session.pop(file_key, None)
            request.session.save()

            return JsonResponse({
                'success': True,
                'inst_id': result.get('inst_id'),
                'message': result.get('message', 'Ticket created successfully.'),
                'department': target_department,
                'requestor': requestor,
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('error', 'Ticket creation failed.'),
            })

    except Exception as e:
        logger.error(f"Error in api_create_eoms_ticket_v2: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ============================================================
# ITSR Ticket Creation API Endpoints (multi-step: credentials → SMS → result)
# ============================================================

# Fixed ITSR ticket parameters
ITSR_TICKET_TITLE = "Service Request - Network Policy"
ITSR_TICKET_DESCRIPTION = "Auto-generated by Netcare"
ITSR_PRODUCT_LINE_ID = "1254515022491552748"  # ISM (網絡支撐)
ITSR_URGENCY = "ZHONG"
ITSR_REQUIREMENT_TYPE = "FEIKAIFAXUQIU"


@require_POST
def api_create_itsr_ticket(request):
    """
    ITSR ticket creation – Step 1: Create session and submit BPM credentials.

    Expects JSON body:
        { "username": "...", "password": "..." }

    The ITSR attachment file path is taken from the Django session (set during
    file processing). Ticket description uses ITSR_TICKET_DESCRIPTION.

    Returns:
        - success + needs_sms=false + result data  → ticket created (no SMS)
        - success + needs_sms=true + session_id     → waiting for SMS code
        - success=false + error                     → login/creation failed
    """
    try:
        data = json.loads(request.body) if request.body else {}
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()

        if not username or not password:
            return JsonResponse({
                'success': False,
                'error': 'Username and password are required.',
            }, status=400)

        # Retrieve ITSR file path from session
        itsr_file_path = request.session.get('itsr_file_path')
        if not itsr_file_path or not os.path.exists(itsr_file_path):
            return JsonResponse({
                'success': False,
                'error': 'Session expired or ITSR file not found. Please re-upload and process the file.',
            }, status=400)

        excel_title = (request.session.get('last_ticket_title') or '').strip()
        itsr_title = excel_title or ITSR_TICKET_TITLE

        attachment_paths = [itsr_file_path]
        orig_path = request.session.get('itsr_original_file_path')
        if orig_path and os.path.exists(orig_path):
            attachment_paths.append(orig_path)

        # 1. Create ITSR session
        session_id = itsr_create_ticket_session(
            title=itsr_title,
            description=ITSR_TICKET_DESCRIPTION,
            product_line_id=ITSR_PRODUCT_LINE_ID,
            urgency=ITSR_URGENCY,
            requirement_type=ITSR_REQUIREMENT_TYPE,
            attachment_files=attachment_paths,  # generated ITSR xlsx + optional user workbook
        )
        logger.info(f"ITSR create session started: {session_id}")

        # Store session_id so SMS endpoint can use it
        request.session['itsr_create_session_id'] = session_id
        request.session.save()

        # 2. Submit credentials (blocks until login completes or times out)
        success, msg = itsr_submit_credentials(session_id, username, password)

        if not success:
            logger.warning(f"ITSR credentials failed for session {session_id}: {msg}")
            return JsonResponse({'success': False, 'error': msg}, status=400)

        # 3a. No SMS required – wait for ticket creation result
        if msg == "NO_SMS_REQUIRED":
            logger.info(f"ITSR session {session_id}: NO_SMS_REQUIRED, waiting for result")
            result = itsr_wait_create_result(session_id, timeout=300)

            if result.success:
                # Mark as created
                request.session['ticket_itsr_created'] = True
                # Clean up session files
                _cleanup_itsr_session_file(itsr_file_path)
                _cleanup_itsr_session_file(request.session.get('itsr_original_file_path'))
                request.session.pop('itsr_file_path', None)
                request.session.pop('itsr_original_file_path', None)
                request.session.pop('itsr_create_session_id', None)
                request.session.save()

                return JsonResponse({
                    'success': True,
                    'needs_sms': False,
                    'bill_code': result.bill_code,
                    'case_id': result.case_id,
                    'subject': result.subject,
                    'message': 'ITSR ticket created successfully.',
                })
            else:
                return JsonResponse({
                    'success': False,
                    'needs_sms': False,
                    'error': result.error or 'ITSR ticket creation failed.',
                })

        # 3b. SMS required
        logger.info(f"ITSR session {session_id}: SMS required")
        return JsonResponse({
            'success': True,
            'needs_sms': True,
            'session_id': session_id,
            'message': 'Login successful. Please check your phone for SMS verification code.',
        })

    except Exception as e:
        logger.error(f"Error in api_create_itsr_ticket: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_POST
def api_submit_itsr_sms(request):
    """
    ITSR ticket creation – Step 2: Submit SMS verification code.

    Expects JSON body:
        { "session_id": "...", "sms_code": "123456" }

    Returns:
        - success + ticket details  → ticket created
        - success=false + error     → creation failed
    """
    try:
        data = json.loads(request.body) if request.body else {}
        session_id = data.get('session_id', '').strip()
        sms_code = data.get('sms_code', '').strip()

        if not session_id:
            return JsonResponse({'success': False, 'error': 'session_id is required.'}, status=400)
        if not sms_code:
            return JsonResponse({'success': False, 'error': 'SMS code is required.'}, status=400)

        # Submit SMS code (blocks until ticket creation completes or times out)
        result = itsr_submit_sms_code(session_id, sms_code)

        if result.success:
            # Mark as created in session
            request.session['ticket_itsr_created'] = True
            # Clean up session files
            itsr_file_path = request.session.get('itsr_file_path')
            _cleanup_itsr_session_file(itsr_file_path)
            _cleanup_itsr_session_file(request.session.get('itsr_original_file_path'))
            request.session.pop('itsr_file_path', None)
            request.session.pop('itsr_original_file_path', None)
            request.session.pop('itsr_create_session_id', None)
            request.session.save()

            return JsonResponse({
                'success': True,
                'bill_code': result.bill_code,
                'case_id': result.case_id,
                'subject': result.subject,
                'message': 'ITSR ticket created successfully.',
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.error or 'ITSR ticket creation failed after SMS.',
            })

    except Exception as e:
        logger.error(f"Error in api_submit_itsr_sms: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@require_GET
def api_check_itsr_create_status(request):
    """
    Check the current status of an ITSR create session.
    Useful for polling while waiting.

    Query param: session_id
    """
    session_id = request.GET.get('session_id', '').strip()
    if not session_id:
        # Fallback: try from Django session
        session_id = request.session.get('itsr_create_session_id', '')

    if not session_id:
        return JsonResponse({'success': False, 'error': 'session_id is required.'}, status=400)

    status = itsr_get_session_status(session_id)
    if status is None:
        return JsonResponse({
            'success': False,
            'error': 'Session not found or expired.',
            'session_id': session_id,
        }, status=404)

    return JsonResponse({
        'success': True,
        'status': status,
        'session_id': session_id,
    })