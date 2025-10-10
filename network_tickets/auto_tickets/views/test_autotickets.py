import re
from netmiko import ConnectHandler
from time import sleep
import openpyxl
import django
import os
import sys
import logging
from datetime import datetime
from pprint import pprint

# Add the parent directory to Python path so we can import network_tickets.settings
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'network_tickets.settings')
django.setup()

# Setup logging
def setup_logging():
    """Setup comprehensive logging for Netmiko sessions"""
    # Create logs directory if it doesn't exist
    log_dir = '/it_network/network_tickets/logs'
    os.makedirs(log_dir, exist_ok=True)
    
    # Create timestamp for log files
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Setup main logger
    logger = logging.getLogger('netmiko_session')
    logger.setLevel(logging.DEBUG)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # File handler for session log
    session_log_file = f'{log_dir}/netmiko_session_{timestamp}.log'
    file_handler = logging.FileHandler(session_log_file)
    file_handler.setLevel(logging.DEBUG)
    
    # Console handler for real-time output
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Setup Netmiko debug logging
    netmiko_logger = logging.getLogger('netmiko')
    netmiko_logger.setLevel(logging.DEBUG)
    
    # Netmiko debug file handler
    debug_log_file = f'{log_dir}/netmiko_debug_{timestamp}.log'
    debug_handler = logging.FileHandler(debug_log_file)
    debug_handler.setLevel(logging.DEBUG)
    debug_handler.setFormatter(formatter)
    netmiko_logger.addHandler(debug_handler)
    
    logger.info(f"Logging initialized - Session log: {session_log_file}")
    logger.info(f"Debug log: {debug_log_file}")
    
    return logger



def auto_tickets_pa_tools(wb, username, password):
    # Initialize logging
    logger = setup_logging()
    logger.info("Starting auto_tickets_pa function")
    res_log = []
    res_log.append('Starting auto_tickets_pa function')
    
    try:
        # wb = openpyxl.load_workbook('/it_network/network_tickets/auto_tickets/sample.xlsx')
        # sheet = wb['Content']
        sheet = wb.active
        logger.info("Excel file loaded successfully")
        res_log.append("Excel file loaded successfully")

        pattern = r'[ ,\n、，]'

        start_row = 4
        end_row = sheet.max_row
        
        if end_row < start_row:
            error_msg = f"No data found in Excel file. Expected data starting from row {start_row}."
            logger.error(error_msg)
            res_log.append(f"ERROR: {error_msg}")
            return res_log

        firewall_PA = {'device_type': 'paloalto_panos',
                   'host': '10.254.0.14',
                   'username': username,
                   'password': password,
                   'session_log': '/it_network/network_tickets/logs/netmiko_session.log',
                   'session_log_file_mode': 'append',
                   'global_delay_factor': 2,
                   }
        
        logger.info(f"Connecting to firewall: {firewall_PA['host']}")
        logger.info(f"Using username: {username}")
        logger.info(f"Using password: {'*' * len(password) if password else 'None'}")
        logger.info(f"Password length: {len(password) if password else 0}")
        res_log.append(f"Connecting to firewall: {firewall_PA['host']}")
        res_log.append(f"Using username: {username}")
        res_log.append(f"Password length: {len(password) if password else 0}")
        
        try:
            net_connect = ConnectHandler(**firewall_PA)
            logger.info("Successfully connected to firewall")
            logger.info(f"Device prompt: {net_connect.find_prompt()}")
            res_log.append(f"Successfully connected to firewall")

        except Exception as e:
            error_msg = f"Failed to connect to firewall: {e}"
            logger.error(error_msg)
            res_log.append(f"ERROR: {error_msg}")
            return res_log


        for row in range(start_row, end_row+1):
            try:
                ticket_number_dic = {}
                sip_dic = {}
                dip_dic = {}
                dport_dic = {}
                protocol_dic = {}

                ticket_number_list = []
                sip_list = []
                dip_list = []
                dport_list = []
                protocol_list = []
                print(f'row: {row}')
                res_log.append(f'Processing row: {row}')
                
                ticket_number = sheet.cell(row=row, column=1).value
                sip = sheet.cell(row=row, column=3).value
                dip = sheet.cell(row=row, column=5).value
                dport = sheet.cell(row=row, column=7).value
                protocol = sheet.cell(row=row, column=6).value
                
                if ticket_number is not None:
                    ticket_number_dic[row] = ticket_number
                else:
                    error_msg = f'ROW {row}: Please provide the ticket number'
                    logger.error(error_msg)
                    res_log.append(f"ERROR: {error_msg}")
                    continue
                

                if protocol is not None: 
                    protocol_dic[row] = protocol
                else:
                    error_msg = f'ROW {row}: Protocol is not defined'
                    logger.error(error_msg)
                    res_log.append(f"ERROR: {error_msg}")
                    continue
                
                if sip is not None:
                    for i in re.split(pattern, sip):
                        try:
                            if re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', i):
                                sip_list.append(i.replace('\u200b', ''))
                        except:
                            error_msg = f'ROW {row}: Source IP is not defined'
                            logger.error(error_msg)
                            res_log.append(f"ERROR: {error_msg}")
                            continue
                    sip_dic[row] = sip_list
                else:
                    error_msg = f'ROW {row}: Source IP is not defined'
                    logger.error(error_msg)
                    res_log.append(f"ERROR: {error_msg}")
                    continue

                if dip is not None:
                    for i in re.split(pattern, dip):
                        try:
                            if re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', i):
                                dip_list.append(i.replace('\u200b', ''))
                        except:
                            error_msg = f'ROW {row}: Destination IP is not defined'
                            logger.error(error_msg)
                            res_log.append(f"ERROR: {error_msg}")
                            continue
                    dip_dic[row] = dip_list
                else:
                    error_msg = f'ROW {row}: Destination IP is not defined'
                    logger.error(error_msg)
                    res_log.append(f"ERROR: {error_msg}")
                    continue
                    
                    
                if dport is not None:
                    for i in re.split(pattern, str(dport)):
                        if i != '':
                            dport_list.append(i.replace('\u200b', ''))
                    dport_dic[row] = dport_list
                else:
                    error_msg = f'ROW {row}: Destination Port is not defined'
                    logger.error(error_msg)
                    res_log.append(f"ERROR: {error_msg}")
                    continue


                sip_zone_dic = {}
                dip_zone_dic = {}

                for sip in sip_dic[row]:
                    try:
                        print(f'sip: {sip}')
                        command = f'test routing fib-lookup virtual-router vr_vsys1 ip {sip}'
                        logger.info(f"Executing command: {command}")
                        check_sip_raw = net_connect.send_command(command)
                        logger.debug(f'check_sip_raw response: {check_sip_raw}')
                        res_log.append(f"Executing command: {command}")
                        print(f'check_sip_raw: {check_sip_raw}')
                        # print(f'check_sip_raw:{check_sip_raw}')

                        # print('-' * 100)
                        interface = re.search(r'interface\s(ethernet\d/\d+),',check_sip_raw).group(1) 
                        command = f'show interface {interface}'
                        logger.info(f"Executing command: {command}")
                        check_int_raw = net_connect.send_command(command)
                        logger.debug(f'check_int_raw response: {check_int_raw}')
                        res_log.append(f"Executing command: {command}")
                        print(f'check_int_raw: {check_int_raw}')
                        zone = re.search(r'Zone:\s(\w+),',check_int_raw).group(1)

                        # print(f'interface:{interface}')
                        # print('-' * 100)
                        logger.info(f"Source IP {sip} mapped to zone {zone}")
                        res_log.append(f"Source IP {sip} mapped to zone {zone}")
                        sip_zone_dic[sip] = zone
                    except Exception as e:
                        error_msg = f"Error processing source IP {sip}: {e}"
                        logger.error(error_msg)
                        res_log.append(f"ERROR: {error_msg}")
                        continue


                for dip in dip_dic[row]:
                    try:
                        command = f'test routing fib-lookup virtual-router vr_vsys1 ip {dip}'
                        logger.info(f"Executing command: {command}")
                        check_dip_raw = net_connect.send_command(command)
                        logger.debug(f'check_dip_raw response: {check_dip_raw}')
                        res_log.append(f"Executing command: {command}")
                        # print(check_dip_raw)
                        interface = re.search(r'interface\s(ethernet\d/\d+),',check_dip_raw).group(1)    
                        
                        command = f'show interface {interface}'
                        logger.info(f"Executing command: {command}")
                        check_int_raw = net_connect.send_command(command)
                        logger.debug(f'check_int_raw response: {check_int_raw}')
                        res_log.append(f"Executing command: {command}")
                        print(f'check_int_raw: {check_int_raw}')
                        zone = re.search(r'Zone:\s(\w+),',check_int_raw).group(1)
                        logger.info(f"Destination IP {dip} mapped to zone {zone}")
                        res_log.append(f"Destination IP {dip} mapped to zone {zone}")
                        dip_zone_dic[dip] = zone
                    except Exception as e:
                        error_msg = f"Error processing destination IP {dip}: {e}"
                        logger.error(error_msg)
                        res_log.append(f"ERROR: {error_msg}")
                        continue

                
                #get source ip and destination ip in reach row, if the zones of source ip and destination ip are different, then add to the set
                sip_ip_set = set()
                dip_ip_set = set()

                for sip, s_zone in sip_zone_dic.items():
                    for dip, d_zone in dip_zone_dic.items():
                        if s_zone != d_zone:
                            print(f'ticket number: {ticket_number_dic[row]}')
                            print(f'source zone: {s_zone}')
                            print(f'destination zone: {d_zone}')
                            print(f'source ip: {sip}')
                            print(f'destination ip: {dip}')
                            print(f'destination port: {dport_dic[row]}')
                            print('-' * 100)
                            res_log.append(f"ticket number: {ticket_number_dic[row]}")
                            res_log.append(f"source zone: {s_zone}")
                            res_log.append(f"destination zone: {d_zone}")
                            res_log.append(f"source ip: {sip}")
                            res_log.append(f"destination ip: {dip}")
                            res_log.append(f"destination port: {dport_dic[row]}")
                            sip_ip_set.add(sip)
                            dip_ip_set.add(dip)
                        else:
                            print(f'{sip} and {dip} belong to the same zone. No action needed')
                            print('-' * 100)
                            res_log.append(f"{sip} and {dip} belong to the same zone. No action needed")

                #skip the current row if there is no action needed   
                if sip_ip_set == set():
                    res_log.append(f"Row {row}: No cross-zone traffic detected, skipping")
                    continue

                print('-' * 100)
                print('Create source ip and destination ip in the firewall')
                res_log.append(f"Create source ip and destination ip in the firewall")
                print('-' * 100)
                
                print(f'source set: {sip_ip_set}')
                print(f'destination set: {dip_ip_set}')
                print('-' * 100)

                src_zone_list = []
                dst_zone_list = []
                for sip in sip_ip_set:
                    src_zone_list.append(sip_zone_dic[sip])
                for dip in dip_ip_set:
                    dst_zone_list.append(dip_zone_dic[dip])
                src_zone_set = set(src_zone_list)
                dst_zone_set = set(dst_zone_list)


                src_zone_ip_dic = {}
                dst_zone_ip_dic = {}
                
                # Initialize dictionaries with empty lists for each zone
                for zone in src_zone_set:
                    src_zone_ip_dic[zone] = []
                for zone in dst_zone_set:
                    dst_zone_ip_dic[zone] = []
                    
                for zone in src_zone_set:
                    for sip in sip_ip_set:
                        if sip_zone_dic[sip] == zone:
                            src_zone_ip_dic[zone].append(sip)
                for zone in dst_zone_set:
                    for dip in dip_ip_set:
                        if dip_zone_dic[dip] == zone:
                            dst_zone_ip_dic[zone].append(dip)
            
                print(f'src_zone_ip_dic: {src_zone_ip_dic}')
                print(f'dst_zone_ip_dic: {dst_zone_ip_dic}')
                print('-' * 100)
                res_log.append(f"src_zone_ip_dic: {src_zone_ip_dic}")
                res_log.append(f"dst_zone_ip_dic: {dst_zone_ip_dic}")
                print('-' * 100)

                print('enter config mode')
                logger.info("Entering configuration mode")
                res_log.append(f"Entering configuration mode")
                out = net_connect.config_mode()
                logger.debug(f"Config mode response: {out}")
                # print(out)
                # print('-' * 100)
                logger.info("Starting address creation in firewall")
                res_log.append(f"Starting address creation in firewall")
        
                
                #create source ip and destination ip in the firewall
                for sip in sip_ip_set:
                    command = f'set address {sip} description generated-by-netcare ip-netmask {sip}'
                    logger.info(f"Creating source address: {command}")
                    res_log.append(f"Creating source address: {command}")
                    try:
                        res = net_connect.send_command(command)
                        logger.info(f"Successfully created address {sip}")
                        print(f"Created address {sip}: {res}")
                        res_log.append(f"Successfully created address {sip}")
                        
                    except Exception as e:
                        error_msg = f"Failed to create address {sip}: {e}"
                        logger.error(error_msg)
                        print(f"Failed to create address {sip}: {e}")
                        res_log.append(f"ERROR: {error_msg}")
                        
                # Create destination addresses  
                for dip in dip_ip_set:
                    command = f'set address {dip} description generated-by-netcare ip-netmask {dip}'
                    logger.info(f"Creating destination address: {command}")
                    res_log.append(f"Creating destination address: {command}")
                    try:
                        res = net_connect.send_command(command)
                        logger.info(f"Successfully created address {dip}")
                        print(f"Created address {dip}: {res}")
                        res_log.append(f"Successfully created address {dip}")
                    
                    except Exception as e:
                        error_msg = f"Failed to create address {dip}: {e}"
                        logger.error(error_msg)
                        print(f"Failed to create address {dip}: {e}")
                        res_log.append(f"ERROR: {error_msg}")
                
                res_log.append(f"Create source address-group and destination address-group in the firewall")
                #create source address-group and destination address-group in the firewall
                print('Create source address-group and destination address-group in the firewall')
                sleep(1)
                # print("Back to operational mode. Prompt:", net_connect.find_prompt())

                # print(type(ticket_number_dic[row]))
                print("Back to operational mode. Prompt:", net_connect.find_prompt())
                
                for zone in src_zone_ip_dic:
                    command = f'edit address-group {ticket_number_dic[row]}-src-row{row}-{zone}'
                    logger.info(f"Creating source address-group: {command}")
                    res_log.append(f"Creating source address-group: {command}")
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully created source address-group: {res}")
                    res_log.append(f"Successfully created source address-group: {res}")
                    print('-' * 100)
                    for sip in src_zone_ip_dic[zone]:
                        command = f'set static {sip}'
                        res = net_connect.send_command(command)
                        logger.info(f"Successfully added static address: {res}")
                        res_log.append(f"Successfully added static address: {res}")
                        print(f'Added static address: {sip}')
                #         print('-' * 100)
                    command = f'set description generated-by-netcare'
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully set description: {res}")
                    res_log.append(f"Successfully set description: {res}")
                    net_connect.send_command('exit')
                    print('-' * 100)    
                
                for zone in dst_zone_ip_dic:
                    command = f'edit address-group {ticket_number_dic[row]}-dst-row{row}-{zone}'
                    logger.info(f"Creating destination address-group: {command}")
                    res_log.append(f"Creating destination address-group: {command}")
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully created destination address-group: {res}")
                    res_log.append(f"Successfully created destination address-group: {res}")
                    print('-' * 100)
                    for dip in dst_zone_ip_dic[zone]:
                        command = f'set static {dip}'
                        res = net_connect.send_command(command)
                        logger.info(f"Successfully added static address: {res}")
                        res_log.append(f"Successfully added static address: {res}")
                        print(f'Added static address: {dip}')
                    command = f'set description generated-by-netcare'
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully set description: {res}")
                    res_log.append(f"Successfully set description: {res}")
                    net_connect.send_command('exit')
                    print('-' * 100)
                
                # # #create service and security rule in the firewall
                print('create service and security rule in the firewall')
                res_log.append(f"create service and security rule in the firewall")
                print("Back to operational mode. Prompt:", net_connect.find_prompt())
                sleep(1)
                # net_connect.send_command('exit')
                # print('create service in the firewall')
                for src_zone in src_zone_set:
                    for dst_zone in dst_zone_set:
                        if src_zone != dst_zone:
                            if protocol_dic[row] == 'tcp' or protocol_dic[row] == 'TCP':
                                for dport in dport_dic[row]:
                                    print(f'dport: {dport}')
                                    print(f'dport_dic[row]: {dport_dic[row]}')
                                    if dport != 'icmp' and dport != 'ICMP':
                                        command = f'set service tcp-{dport}-netcare description generated-by-netcare protocol tcp port {dport}'
                                        logger.info(f"Creating service: {command}")
                                        res_log.append(f"Creating service: {command}")
                                        res = net_connect.send_command(command)
                                        logger.info(f"Successfully created service: {res}")
                                        res_log.append(f"Successfully created service: {res}")

                                        command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-tcp' + dport + '-row' + str(row) + '-' + src_zone + '-' + dst_zone + '-netcare', 
                                        'set source ' + ticket_number_dic[row] + '-src-row' + str(row) +'-' + src_zone + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + '-' + dst_zone + ' from ' + src_zone + ' to ' + dst_zone + ' service ' + 'tcp-' + dport + '-netcare',
                                        'set description generated-by-netcare','set application any','set action allow','exit', 'exit', 'exit', 'move rulebase security rules ' + ticket_number_dic[row] + '-tcp' + dport + '-row' + str(row) + '-' + src_zone + '-' + dst_zone + '-netcare top']
                                        
                                        res_log.append(f"Creating security rule: {command_list}")

                                        res = net_connect.send_config_set(command_list, exit_config_mode=False)
                                        sleep(1)
                                        logger.info(f"Successfully created security rule: {res}")

                                        # print("Back to operational mode. Prompt:", net_connect.find_prompt())

                                        
                                    else:
                                        command_list = ['edit rulebase security rules ' + 'icmp' + '-netcare', 
                                        'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + '-' + src_zone + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + '-' + dst_zone + ' from any to any service application-default application icmp ',
                                        'set description generated-by-netcare','set action allow','exit', 'exit', 'exit']
                                        res_log.append(f"Creating security rule: {command_list}")
                                    
                                        res = net_connect.send_config_set(command_list, exit_config_mode=False)
                                        sleep(1)
                                        logger.info(f"Successfully created security rule: {res}")

                            elif protocol_dic[row] == 'udp' or protocol_dic[row] == 'UDP':
                                for dport in dport_dic[row]:
                                    if dport != 'icmp' and dport != 'ICMP':
                                        command = f'set service udp-{dport}-netcare description generated-by-netcare protocol udp port {dport}'
                                        logger.info(f"Creating service: {command}")
                                        res_log.append(f"Creating service: {command}")
                                        res = net_connect.send_command(command)
                                        logger.info(f"Successfully created service: {res}")
                                        res_log.append(f"Successfully created service: {res}")
                                        
                                        command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-udp' + dport + '-row' + str(row) + '-' + src_zone + '-' + dst_zone + '-netcare', 
                                        'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + '-' + src_zone + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + '-' + dst_zone + ' from ' + src_zone + ' to ' + dst_zone + ' service ' + 'udp-' + dport + '-netcare',
                                        'set description generated-by-netcare','set application any','set action allow','exit', 'exit', 'exit', 'move rulebase security rules ' + ticket_number_dic[row] + '-udp' + dport + '-row' + str(row) + '-' + src_zone + '-' + dst_zone + '-netcare top']
                                        res_log.append(f"Creating security rule: {command_list}")
                                        
                                        res = net_connect.send_config_set(command_list, exit_config_mode=False)
                                        sleep(1)
                                        logger.info(f"Successfully created security rule: {res}")
                                        res_log.append(f"Successfully created security rule: {res}")
                                    else:
                                        command_list = ['edit rulebase security rules ' + 'icmp' + '-' + 'netcare', 
                                        'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + '-' + src_zone + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + '-' + dst_zone + ' from any to any service application-default application icmp ',
                                        'set description generated-by-netcare','set action allow','exit', 'exit', 'exit']
                                        res_log.append(f"Creating security rule: {command_list}")
                                        res = net_connect.send_config_set(command_list, exit_config_mode=False)
                                        sleep(1)
                                        logger.info(f"Successfully created security rule: {res}")
                                        res_log.append(f"Successfully created security rule: {res}")
                            elif protocol_dic[row] == 'icmp' or protocol_dic[row] == 'ICMP':
                                command_list = ['edit rulebase security rules ' + 'icmp' + '-' + 'netcare', 
                                'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + '-' + src_zone + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + '-' + dst_zone + ' from any to any service application-default application icmp ',
                                'set description generated-by-netcare','set action allow','exit', 'exit', 'exit']
                                res_log.append(f"Creating security rule: {command_list}")

                                res = net_connect.send_config_set(command_list, exit_config_mode=False)
                                sleep(1)
                                logger.info(f"Successfully created security rule: {res}")
                                res_log.append(f"Successfully created security rule: {res}")
                            else:
                                error_msg = 'Pls check the protocol and port in the excel file.'
                                logger.error(error_msg)
                                res_log.append(f"ERROR: {error_msg}")

                
                print('I am here')
                res_log.append(f"Exiting configuration mode")
                logger.info("Exiting configuration mode")
                # print("Back to operational mode. Prompt:", net_connect.find_prompt())
                net_connect.exit_config_mode()   
            
            except Exception as e:
                error_msg = f"Error processing row {row}: {e}"
                logger.error(error_msg)
                res_log.append(f"ERROR: {error_msg}")
                continue
 
         #check if there is any job running before validate
        while True:
            command = 'show jobs all'
            output = net_connect.send_command(command)
            print(f'Show jobs status: {output}')
            logger.info(f"Show jobs status: {output}")
            res_log.append(f"Show jobs status: {output}")
            if 'ACT' in output:
                sleep(60)
            else:
                break
        
        #validate job
        net_connect.config_mode()
        logger.info(f"Back to operational mode. Prompt: {net_connect.find_prompt()}")
        command ='validate partial admin ' + username
        output = net_connect.send_command(command)
        net_connect.send_command('exit', expect_string=r'>')
        logger.info(f"Back to check mode. Prompt: {net_connect.find_prompt()}")
        print(f'Validate status: {output}')
        logger.info(f"Validate status: {output}")
        res_log.append(f"Validate status: {output}")
        validate_job_id = re.search(r'jobid\s+(\d+)', output).group(1)
        print(f'Validate job id: {validate_job_id}')
        logger.info(f"Validate job id: {validate_job_id}")
        res_log.append(f"Validate job id: {validate_job_id}")
        
        
        while True:
            command = f'show jobs id {validate_job_id}'
            output = net_connect.send_command(command)
            print(f'Validate status: {output}')
            logger.info(f"Validate status: {output}")
            res_log.append(f"Validate status: {output}")
            if 'FIN' in output and 'OK' in output:
                res_log.append("Validation completed successfully")
                break   # next step is commit
            elif 'FAIL' in output:
                error_msg = "Validation failed"
                logger.error(error_msg)
                res_log.append(f"ERROR: {error_msg}")
                net_connect.disconnect()
                res_log.append(f"Disconnecting from firewall")
                return res_log  
            else:
                res_log.append("Validation still in progress, waiting...")
                sleep(60)
        print('Validate status: FIN OK')
        logger.info("Validate status: FIN OK")
        res_log.append(f"Validate status: FIN OK")

        #check if there is any job running before commit
        while True:
            command = 'show jobs all'
            output = net_connect.send_command(command)
            print(f'Show jobs status: {output}')
            logger.info(f"Show jobs status: {output}")
            res_log.append(f"Show jobs status: {output}")
            if 'ACT' in output:
                sleep(60)
            else:
                break
        net_connect.config_mode()
        logger.info(f"Back to operational mode. Prompt: {net_connect.find_prompt()}")
        commit_command = 'commit partial admin ' + username + ' description commit-by-netcare'
        commit_output = net_connect.send_command_timing(commit_command,read_timeout=0)
        print(f"Successfully committed: {commit_output}")
        logger.info(f"Successfully committed: {commit_output}")
        res_log.append(f"{commit_output}") 
        logger.info("Disconnecting from firewall")
        net_connect.disconnect()
        res_log.append(f"Disconnecting from firewall")
        logger.info("Successfully disconnected from firewall")
        logger.info("auto_tickets_pa function completed")
        return res_log
    
    except Exception as e:
        error_msg = f"Critical error in auto_tickets_pa_tools: {e}"
        logger.error(error_msg)
        res_log.append(f"ERROR: {error_msg}")
        try:
            net_connect.disconnect()
            res_log.append("Disconnected from firewall after error")
        except:
            pass
        return res_log




wb = openpyxl.load_workbook('/it_network/network_tickets/auto_tickets/sample (3).xlsx')
auto_tickets_pa_tools(wb, 'ZhengCheng', 'Cmhk@941')