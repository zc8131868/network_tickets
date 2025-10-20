import sys
from pysqlite3 import dbapi2 as sqlite3
sys.modules['sqlite3'] = sqlite3

import chromadb
from chromadb import PersistentClient
from agents import set_default_openai_key, set_tracing_disabled, AsyncOpenAI

import os
import glob
import time
from datetime import datetime
from pathlib import Path
import openpyxl
import asyncio


FILES_DIR = '/it_network/network_tickets/auto_tickets/ai_tools/files_dir'
DB_PATH = '/it_network/network_tickets/auto_tickets/ai_tools/chromadb.db'

CHUNK_SIZE = 500
CHUNK_OVERLAP = 150

BATCH_SIZE = 100
MODEL = 'text-embedding-3-small'
COLLECTION_NAME = 'NetCare_RAG'

# Try to get API key from environment variable
aihubmix_apikey = os.environ.get("AIHUBMIX")

# If not in environment, try to load from .env file
if not aihubmix_apikey:
    env_file = Path(__file__).parent / '.env'
    if env_file.exists():
        with open(env_file) as f:
            for line in f:
                if line.startswith('AIHUBMIX='):
                    aihubmix_apikey = line.strip().split('=', 1)[1].strip('"').strip("'")
                    break

if not aihubmix_apikey:
    raise ValueError("AIHUBMIX API key not found. Please set AIHUBMIX environment variable or create a .env file with AIHUBMIX=your-api-key")

set_default_openai_key(aihubmix_apikey)

set_tracing_disabled(True)


openai_client = AsyncOpenAI(
    api_key=aihubmix_apikey,
    base_url="https://aihubmix.com/v1/",
)

def init_clients():
    # openai_client = OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
    chroma_client = PersistentClient(path=DB_PATH)
    collection = chroma_client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={'description': 'NetCare_RAG_chunks'},
        )
    return openai_client, collection


def clear_collection(collection):
    #get the ids of allthe documents in the collection
    all_ids = collection.get(include=[])['ids']
    #delete the documents with the ids
    collection.delete(ids=all_ids)
    print(f"Deleted {len(all_ids)} documents from the collection")


def find_docs():
    md_files = glob.glob(os.path.join(FILES_DIR, '*.md'))
    xlsx_files = glob.glob(os.path.join(FILES_DIR, '*.xlsx'))
    file_list = md_files + xlsx_files

    if not file_list:
        print("No files found in the directory")
        return
    else:
        print(f"Found {len(file_list)} files in the directory ({len(md_files)} .md, {len(xlsx_files)} .xlsx)")
        return file_list


def extract_text_from_excel(file_path):
    """Extract text content from Excel file"""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    text_parts = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        text_parts.append(f"Sheet: {sheet_name}\n")
        
        # Get all rows with values
        for row in ws.iter_rows(values_only=True):
            # Filter out None values and convert to strings
            row_values = [str(cell) if cell is not None else '' for cell in row]
            # Only add non-empty rows
            if any(row_values):
                text_parts.append(' | '.join(row_values))
        
        text_parts.append('\n')
    
    return '\n'.join(text_parts)


def generate_chunks(text):

    if len(text) < CHUNK_SIZE:
        return [text]
    
    chunks = []

    step = CHUNK_SIZE - CHUNK_OVERLAP

    for i in range(0, len(text), step):
        start = i
        end = min(start + CHUNK_SIZE, len(text))
        chunks.append(text[start:end])
        if end == len(text):
            break
    return chunks


async def create_embedding(openai_client, text):
    response = await openai_client.embeddings.create(
        input=[text],
        model=MODEL
    )
    return response.data[0].embedding

def write_batch_to_collection(collection, ids_batch, embeddings_batch, docs_batch, metas_batch, batch_count):

    collection.add(
            embeddings=embeddings_batch,
            documents=docs_batch,
            metadatas=metas_batch,
            ids=ids_batch
    )
    batch_count += 1
    print(f"Added {len(ids_batch)} documents to the collection in batch {batch_count}")
    return batch_count


async def process_files(openai_client, collection, file_list):
    ids_batch = []
    embeddings_batch = []
    docs_batch = []
    metas_batch = []
    batch_count = 0
    created_at = datetime.now().isoformat()
    total_chunks = 0
    for file_idx, file_path in enumerate(file_list,1):
        try:
            filename = os.path.basename(file_path)
            ext = os.path.splitext(filename)[1].lstrip('.')
            
            print(f"Processing file {file_idx} of {len(file_list)}: {filename}")
            
            # Extract text based on file type
            if ext == 'xlsx':
                text = extract_text_from_excel(file_path)
            else:  # markdown or other text files
                with open(file_path, encoding='utf-8') as f:
                    text = f.read()
            
            chunks = generate_chunks(text)
            print(f"Generated {len(chunks)} chunks from {filename}")
            
            for idx, chunk in enumerate(chunks):
                emb = await create_embedding(openai_client, chunk)
                chunk_id = f"{filename}-{idx}"
                metadata = {
                    'filename': filename,
                    'chunk_idx': idx,
                    'type': ext,
                    'created_at': created_at
                }

                ids_batch.append(chunk_id)
                embeddings_batch.append(emb)
                docs_batch.append(chunk)
                metas_batch.append(metadata)
                total_chunks += 1

                if len(ids_batch) >= BATCH_SIZE:
                    batch_count = write_batch_to_collection(collection, ids_batch, embeddings_batch, docs_batch, metas_batch, batch_count)
                    ids_batch = []
                    embeddings_batch = []
                    docs_batch = []
                    metas_batch = []
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            continue
    if ids_batch:
        batch_count = write_batch_to_collection(collection, ids_batch, embeddings_batch, docs_batch, metas_batch, batch_count)
    print(f"Total chunks processed: {total_chunks}")
    print(f"Total batches processed: {batch_count}")
    return total_chunks
                        

if __name__ == "__main__":
    Path(DB_PATH).mkdir(parents=True, exist_ok=True)
    try:
        start_time = time.time()
        openai_client, collection = init_clients()
        file_list = find_docs()
        if not file_list:
            exit(1)

        total_chunks = asyncio.run(process_files(openai_client, collection, file_list))
        end_time = time.time()
        print(f"Total time taken: {end_time - start_time} seconds")
    except Exception as e:
        print(f"Error: {e}")    