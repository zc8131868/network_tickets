import os
import time
import requests
import openpyxl
import re
from datetime import datetime
from datetime import timedelta

# Phone number validation patterns
# Supports: 11-digit mainland China numbers, +852/HK 8-digit, +86/86 11-digit
PHONE_PATTERN = re.compile(
    r'^('
    r'\+?852\d{8}|'      # Hong Kong: +85212345678 or 85212345678
    r'\+?86\d{11}|'      # China with country code: +8613800138000 or 8613800138000
    r'\d{11}'            # Mainland China: 13800138000
    r')$'
)


class VPNApiClient:
    """VPN API client with automatic token refresh"""
    
    def __init__(self):
        self.endpoint = 'https://vpn.hk.chinamobile.com:8443'
        self.access_key_id = 'kGDarBmLTGmYzVtlPxpH'
        self.access_key_secret = 'wvYDfEVOaFlKUgnbEmCJoanojdiTBTjCXmvnLKde'
        self.token = None
        self.token_expires_at = 0
        self.department_id = None
    
    def _is_token_valid(self) -> bool:
        """Check if current token is still valid (with 60s buffer)"""
        return self.token is not None and time.time() < (self.token_expires_at - 60)
    
    def _get_token(self) -> str:
        """Get a valid token, refreshing if necessary"""
        if self._is_token_valid():
            return self.token
        
        url = f'{self.endpoint}/api/open/v1/token'
        payload = {
            'access_key_id': self.access_key_id,
            'access_key_secret': self.access_key_secret
        }
        
        try:
            response = requests.post(url, json=payload)
            response.raise_for_status()
            res = response.json()
            print(f'Token response: {res}')
            
            if 'data' not in res:
                raise Exception(f"Unexpected token response: {res}")
            
            self.token = res['data']['access_token']
            expires_in = res['data'].get('expires_in', 7200)
            self.token_expires_at = time.time() + expires_in
            print(f'✅ Token refreshed, expires in {expires_in} seconds')
            return self.token
        except requests.exceptions.RequestException as e:
            print(f'Error getting token: {e}')
            raise
        except KeyError as e:
            print(f'Error parsing token response: {e}')
            raise
    
    def _get_headers(self) -> dict:
        """Get headers with valid token"""
        return {'Authorization': self._get_token()}
    
    def get_department_id(self, name: str = "Vendor") -> str:
        """Get department ID by name"""
        if self.department_id is not None:
            return self.department_id
        
        url = f'{self.endpoint}/api/open/v1/department/get_id'
        payload = {'name': name}
        
        try:
            response = requests.post(url, json=payload, headers=self._get_headers())
            response.raise_for_status()
            self.department_id = response.json()['data']['id']
            print(f'department_id: {self.department_id}')
            return self.department_id
        except requests.exceptions.RequestException as e:
            print(f'Error getting department ID: {e}')
            raise
    
    def get_manager_id(self, manager_email: str) -> str:
        """Get manager ID by email"""
        url = f'{self.endpoint}/api/open/v1/user/get_id'
        payload = {'email': manager_email}
        
        try:
            response = requests.post(url, json=payload, headers=self._get_headers())
            response.raise_for_status()
            res = response.json()
            return res['data']['id']
        except requests.exceptions.RequestException as e:
            print(f'Error getting manager ID: {e}')
            return None
    
    def create_user(self, vendor_name: str, vendor_email: str, phone_number: str, manager_email: str) -> dict:
        """Create a VPN user account"""
        now = datetime.now()
        formatted_date_after1year = (now + timedelta(days=365)).strftime("%Y-%m-%d")
        
        manager_id = self.get_manager_id(manager_email)
        if manager_id is None:
            print(f'Failed to get manager id for {manager_email}')
            return {'success': False, 'user_id': None, 'error': f'{vendor_name}: Failed to get manager ID for {manager_email}'}
        
        department_id = self.get_department_id()
        
        url = f'{self.endpoint}/api/open/v1/user/create'
        payload = {
            'full_name': vendor_name,
            'email': vendor_email,
            'mobile': phone_number,
            'manager_ids': [manager_id],
            'department_id': department_id,
            'invite_type': 3,
            'expired_at': formatted_date_after1year,
        }
        
        try:
            response = requests.post(url, json=payload, headers=self._get_headers())
            response.raise_for_status()
            result = response.json()
            
            if 'data' in result:
                if isinstance(result['data'], dict) and 'result' in result['data']:
                    if result['data']['result'] == 'success':
                        return {'success': True, 'user_id': result['data']['id'], 'error': None}
                    else:
                        error_msg = result.get('message', 'Unknown error')
                        print(f'{vendor_name}: {error_msg}')
                        return {'success': False, 'user_id': None, 'error': f'{vendor_name}: {error_msg}'}
                else:
                    error_msg = f'Unexpected response structure: {result}'
                    print(error_msg)
                    return {'success': False, 'user_id': None, 'error': f'{vendor_name}: {error_msg}'}
            else:
                if 'message' in result:
                    error_msg = f'{vendor_name}: {result["message"]}'
                    print(error_msg)
                    return {'success': False, 'user_id': None, 'error': error_msg}
                else:
                    error_msg = f'Unexpected response format: {result}'
                    print(error_msg)
                    return {'success': False, 'user_id': None, 'error': f'{vendor_name}: {error_msg}'}
        except requests.exceptions.RequestException as e:
            error_msg = f'{vendor_name}: {e}'
            print(error_msg)
            return {'success': False, 'user_id': None, 'error': error_msg}
        except KeyError as e:
            error_msg = f'KeyError: Missing key {e} in response: {result}'
            print(error_msg)
            return {'success': False, 'user_id': None, 'error': f'{vendor_name}: {error_msg}'}
    
    def get_user_id(self, identifier: str) -> str:
        """Get user ID by email or name"""
        url = f'{self.endpoint}/api/open/v1/user/get_id'
        payload = {'email': identifier}
        
        try:
            response = requests.post(url, json=payload, headers=self._get_headers())
            response.raise_for_status()
            res = response.json()
            return res['data']['id']
        except requests.exceptions.RequestException as e:
            print(f'Error getting user ID for {identifier}: {e}')
            return None
        except KeyError:
            print(f'User not found: {identifier}')
            return None
    
    def create_resource(self, ticket_number: str, dip_list: list, dport_list: list, protocol_list: list) -> dict:
        """Create a VPN ACL resource"""
        url = f'{self.endpoint}/api/open/v1/vpn/acl/resource/create'
        payload = {
            'key': ticket_number,
            'type': 'ip',
            'protocols': protocol_list,
            'addrs': dip_list,
            'ports': dport_list,
        }
        
        try:
            response = requests.post(url, json=payload, headers=self._get_headers())
            response.raise_for_status()
            result = response.json()
            if result['code'] == 0:
                print(f'Resource creation for {ticket_number} successful: {result}')
                return {'success': True, 'resource_id': result["data"]["resource_id"], 'error': None}
            else:
                error_msg = result.get('message', 'Unknown error')
                print(f'{ticket_number}: Resource creation failed - {error_msg}')
                return {'success': False, 'resource_id': None, 'error': f'{ticket_number}: Resource creation failed - {error_msg}'}
        except requests.exceptions.RequestException as e:
            error_msg = f'Resource creation for {ticket_number} failed: {e}'
            print(error_msg)
            return {'success': False, 'resource_id': None, 'error': error_msg}
        except KeyError as e:
            error_msg = f'Resource creation for {ticket_number} failed: Missing key {e} in response: {result}'
            print(error_msg)
            return {'success': False, 'resource_id': None, 'error': error_msg}
        except Exception as e:
            error_msg = f'Resource creation for {ticket_number} failed: {e}'
            print(error_msg)
            return {'success': False, 'resource_id': None, 'error': error_msg}
    
    def create_policy(self, ticket_number: str, resource_id_list: list, user_id_list: list) -> dict:
        """Create a VPN ACL policy"""
        url = f'{self.endpoint}/api/open/v1/vpn/acl/policy/create'
        payload = {
            'key': ticket_number,
            'action': 1,
            'priority': 0,
            'resource_ids': resource_id_list,
            'user_ids': user_id_list,
        }
        
        try:
            response = requests.post(url, json=payload, headers=self._get_headers())
            response.raise_for_status()
            result = response.json()
            if result['code'] == 0:
                print(f'Policy creation for {ticket_number} successful: {result}')
                return {'success': True, 'policy_id': result["data"]["id"], 'error': None}
            else:
                error_msg = result.get('message', 'Unknown error')
                print(f'{ticket_number}: Policy creation failed - {error_msg}')
                return {'success': False, 'policy_id': None, 'error': f'{ticket_number}: Policy creation failed - {error_msg}'}
        except requests.exceptions.RequestException as e:
            error_msg = f'Policy creation for {ticket_number} failed: {e}'
            print(error_msg)
            return {'success': False, 'policy_id': None, 'error': error_msg}
        except KeyError as e:
            error_msg = f'Policy creation for {ticket_number} failed: Missing key {e} in response'
            print(error_msg)
            return {'success': False, 'policy_id': None, 'error': error_msg}
        except Exception as e:
            error_msg = f'Policy creation for {ticket_number} failed: {e}'
            print(error_msg)
            return {'success': False, 'policy_id': None, 'error': error_msg}


# Global client instance (lazy initialization - no API calls until first use)
_vpn_client = None

def get_vpn_client() -> VPNApiClient:
    """Get or create the VPN API client"""
    global _vpn_client
    if _vpn_client is None:
        _vpn_client = VPNApiClient()
    return _vpn_client



def create_vpn_user_tool(wb):
    # Get VPN API client (with auto token refresh)
    vpn_client = get_vpn_client()
    
    sheet = wb.active

    start_row = 4
    end_row = sheet.max_row
    print(f'end_row: {end_row}')

    if end_row < start_row:
        error_msg = f"No data found in Excel file. Expected data starting from row {start_row}."
        print(error_msg)
        exit(1)

    ticket_number_dic = {}
    manager_email_dic = {}
    vendor_name_dic = {}
    vendor_email_dic = {}
    phone_number_dic = {}
    validation_errors = []

    for row in range(start_row, end_row + 1):
        print(f'row: {row}')

        ticket_number = sheet.cell(row=row, column=1).value
        manager_email = sheet.cell(row=row, column=2).value
        vendor_name = sheet.cell(row=row, column=3).value
        vendor_email = sheet.cell(row=row, column=4).value
        phone_number = sheet.cell(row=row, column=5).value

        # Skip empty rows
        if not any([ticket_number, manager_email, vendor_name, vendor_email, phone_number]):
            continue

        row_errors = []
        
        if ticket_number is not None:
            ticket_number_dic[row] = str(ticket_number)
        else:
            error_msg = f'Row {row}: Ticket number is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if phone_number is not None:
            # Clean phone number (remove spaces, dashes, parentheses)
            cleaned_phone = re.sub(r'[\s\-\(\)]', '', str(phone_number))
            
            # Validate phone number format
            if PHONE_PATTERN.match(cleaned_phone):
                phone_number_dic[row] = cleaned_phone
            else:
                vendor_display = vendor_name if vendor_name else f'Row {row}'
                error_msg = f'{vendor_display}: 手机号格式错误 (Invalid phone number format: {phone_number}). Expected: 11-digit mainland China number, +852 + 8-digit HK number, or +86 + 11-digit China number'
                print(error_msg)
                row_errors.append(error_msg)
        else:
            vendor_display = vendor_name if vendor_name else f'Row {row}'
            error_msg = f'{vendor_display}: Phone number is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if manager_email is not None:
            if re.search(r'^[a-zA-Z0-9._%+-]+@hk\.chinamobile\.com$', manager_email):
                manager_email_dic[row] = manager_email
            else:
                error_msg = f'Row {row}: Invalid email format ({manager_email})'
                print(error_msg)
                row_errors.append(error_msg)
        else:
            error_msg = f'Row {row}: Manager email is not found'
            print(error_msg)
            row_errors.append(error_msg)


        if vendor_name is not None:
            vendor_name_dic[row] = vendor_name.replace(' ', '')
        else:
            error_msg = f'Row {row}: Vendor name is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if vendor_email is not None:
            if re.search(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$', vendor_email):
                vendor_email_dic[row] = vendor_email
            else:
                vendor_display = vendor_name if vendor_name else f'Row {row}'
                error_msg = f'{vendor_display}: Invalid email format ({vendor_email})'
                print(error_msg)
                row_errors.append(error_msg)
        else:
            vendor_display = vendor_name if vendor_name else f'Row {row}'
            error_msg = f'{vendor_display}: Vendor email is not found'
            print(error_msg)
            row_errors.append(error_msg)
        
        # Only add row to processing if no validation errors
        if row_errors:
            validation_errors.extend(row_errors)
            # Remove row from dictionaries if it was partially added
            ticket_number_dic.pop(row, None)
            manager_email_dic.pop(row, None)
            vendor_name_dic.pop(row, None)
            vendor_email_dic.pop(row, None)
            phone_number_dic.pop(row, None)

    res = {}
    errors = []
    
    # Add validation errors first
    errors.extend(validation_errors)
    
    # Process valid rows
    for row in range(start_row, end_row + 1):
        if row not in vendor_name_dic:
            continue  # Skip rows with validation errors
        
        result = vpn_client.create_user(vendor_name_dic[row], vendor_email_dic[row], phone_number_dic[row], manager_email_dic[row])
        if result['success']:
            print(f'Successful to create an account for {vendor_name_dic[row]}')
            res[vendor_name_dic[row]] = result['user_id']
        else:
            error_msg = result['error'] or f'Failed to create an account for {vendor_name_dic[row]}'
            print(error_msg)
            errors.append(error_msg)
    
    return {'success': res, 'errors': errors}


if __name__ == '__main__':
    # Get the directory where this script is located
    script_dir = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(script_dir, 'VPN_Account_Ticket_Sample_test.xlsx')
    wb = openpyxl.load_workbook(file_path)
    res = create_vpn_user_tool(wb)
    print(f'res: {res}')