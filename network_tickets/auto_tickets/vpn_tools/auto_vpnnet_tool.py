import openpyxl
import re

# Import the shared VPN API client with automatic token refresh
from auto_tickets.vpn_tools.create_user_tool import get_vpn_client

def create_vpn_access_policy_tool(wb):
    # Get VPN API client (with auto token refresh)
    vpn_client = get_vpn_client()
    
    sheet = wb.active
    pattern = r'[ ,\n、，]'
    start_row = 4
    end_row = sheet.max_row

    if end_row < start_row:
        error_msg = f"No data found in Excel file. Expected data starting from row {start_row}."
        print(error_msg)
        return {'success': [], 'errors': [error_msg]}

    ticket_number_dic = {}
    dip_dic = {}
    dport_dic = {}
    protocol_dic = {}
    vendor_name_dic = {}
    validation_errors = []

    for row in range(start_row, end_row + 1):
        dip_list = []
        dport_list = []
        vendor_name_list = []
        protocol_list = []

        print(f'row: {row}')

        ticket_number = sheet.cell(row=row, column=1).value
        dip = sheet.cell(row=row, column=2).value
        protocol = (sheet.cell(row=row, column=3).value)
        dport = sheet.cell(row=row, column=4).value
        vendor_name = sheet.cell(row=row, column=5).value

        # Skip empty rows
        if not any([ticket_number, dip, protocol, dport, vendor_name]):
            continue

        row_errors = []

        if ticket_number is not None:
            ticket_number_dic[row] = str(ticket_number)
        else:
            error_msg = f'Row {row}: Ticket number is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if dip is not None:
            for i in re.split(pattern, dip):
                try:
                    # Clean IP address: remove all whitespace and zero-width characters
                    cleaned_ip = re.sub(r'[\s\u200b]+', '', i)
                    if cleaned_ip and re.search(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', cleaned_ip):
                        dip_list.append(cleaned_ip)
                    elif cleaned_ip:
                        error_msg = f'Row {row}: Invalid IP address: {cleaned_ip}'
                        print(error_msg)
                        row_errors.append(error_msg)
                except Exception as e:
                    error_msg = f'Row {row}: Error processing IP {i}: {e}'
                    print(error_msg)
                    row_errors.append(error_msg)
            if dip_list:
                dip_dic[row] = dip_list
            elif dip is not None:
                error_msg = f'Row {row}: No valid IP addresses found'
                row_errors.append(error_msg)
        else:
            error_msg = f'Row {row}: Destination IP is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if dport is not None:
            for i in re.split(pattern, str(dport)):
                try:
                    # Clean port number: remove all whitespace and zero-width characters
                    cleaned_dport = re.sub(r'[\s\u200b]+', '', i)
                    if cleaned_dport:
                        dport_list.append(cleaned_dport)
                except Exception as e:
                    error_msg = f'Row {row}: Error processing port {i}: {e}'
                    print(error_msg)
                    row_errors.append(error_msg)
            if dport_list:
                dport_dic[row] = dport_list
        else:
            error_msg = f'Row {row}: Destination port is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if protocol is not None:
            for i in re.split(pattern, protocol.lower()):
                try:
                    cleaned_protocol = re.sub(r'[\s\u200b]+', '', i)
                    if cleaned_protocol:
                        protocol_list.append(cleaned_protocol)
                except Exception as e:
                    error_msg = f'Row {row}: Error processing protocol {i}: {e}'
                    print(error_msg)
                    row_errors.append(error_msg)
            if protocol_list:
                protocol_dic[row] = protocol_list
        else:
            error_msg = f'Row {row}: Protocol is not found'
            print(error_msg)
            row_errors.append(error_msg)

        if vendor_name is not None:
            for i in re.split(pattern, vendor_name):
                try:
                    cleaned_vendor_name = re.sub(r'[\s\u200b]+', '', i)
                    if cleaned_vendor_name:
                        vendor_name_list.append(cleaned_vendor_name)
                except Exception as e:
                    error_msg = f'Row {row}: Error processing vendor name {i}: {e}'
                    print(error_msg)
                    row_errors.append(error_msg)
            if vendor_name_list:
                vendor_name_dic[row] = vendor_name_list
        else:
            error_msg = f'Row {row}: Vendor name is not found'
            print(error_msg)
            row_errors.append(error_msg)

        # If row has errors, remove it from processing dictionaries
        if row_errors:
            validation_errors.extend(row_errors)
            ticket_number_dic.pop(row, None)
            dip_dic.pop(row, None)
            dport_dic.pop(row, None)
            protocol_dic.pop(row, None)
            vendor_name_dic.pop(row, None)

    # Process valid rows
    success_results = []
    errors = []
    errors.extend(validation_errors)
    
    for row in range(start_row, end_row + 1):
        if row not in ticket_number_dic:
            continue  # Skip rows with validation errors

        user_id_list = []
        for vendor_name in vendor_name_dic[row]:
            user_id = vpn_client.get_user_id(vendor_name)
            if user_id:
                user_id_list.append(user_id)
            else:
                error_msg = f'Row {row}: Failed to get user id for vendor {vendor_name}'
                print(error_msg)
                errors.append(error_msg)
                break  # Skip this row if we can't get user IDs
        
        if not user_id_list:
            continue  # Skip this row if no valid user IDs

        ticket_number = ticket_number_dic[row] + "_" + str(row)

        resource_result = vpn_client.create_resource(ticket_number, dip_dic[row], dport_dic[row], protocol_dic[row])
        if resource_result['success']:
            print(f'Successful to create a resource policy for {ticket_number}')
            resource_id = resource_result['resource_id']
            
            policy_result = vpn_client.create_policy(ticket_number, [resource_id], user_id_list)
            if policy_result['success']:
                print(f'Successful to create a policy for {ticket_number}')
                success_results.append({
                    'ticket_number': ticket_number,
                    'resource_id': resource_id,
                    'policy_id': policy_result['policy_id'],
                    'row': row
                })
            else:
                error_msg = policy_result['error'] or f'Row {row}: Failed to create policy for {ticket_number}'
                print(error_msg)
                errors.append(error_msg)
        else:
            error_msg = resource_result['error'] or f'Row {row}: Failed to create resource for {ticket_number}'
            print(error_msg)
            errors.append(error_msg)

    return {'success': success_results, 'errors': errors}