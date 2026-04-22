import unittest
from unittest.mock import patch
import os
import sys
import types

import openpyxl

sys.path.insert(0, "/it_network/network_tickets")
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "network_tickets.settings")

import django

django.setup()

fake_eoms = types.ModuleType("auto_tickets.views.ITSR_Tools.eoms_automation_2")
fake_eoms.create_ticket = lambda *args, **kwargs: None
sys.modules["auto_tickets.views.ITSR_Tools.eoms_automation_2"] = fake_eoms

fake_itsr_create = types.ModuleType("auto_tickets.views.ITSR_Tools.itsr_create")
fake_itsr_create.create_ticket_session = lambda *args, **kwargs: None
fake_itsr_create.submit_credentials = lambda *args, **kwargs: None
fake_itsr_create.submit_sms_code = lambda *args, **kwargs: None
fake_itsr_create.wait_create_result = lambda *args, **kwargs: None
fake_itsr_create.cancel_session = lambda *args, **kwargs: None
fake_itsr_create.get_session_status = lambda *args, **kwargs: None
sys.modules["auto_tickets.views.ITSR_Tools.itsr_create"] = fake_itsr_create

from auto_tickets.views.multi_split import _process_itsr_file


def _build_sheet():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=9).value = "Ticket Title"
    return sheet


def _set_itsr_row(sheet, row_num, source_name, source_ip, dest_name, dest_ip, port, protocol, requestor=""):
    sheet.cell(row=row_num, column=2).value = source_name
    sheet.cell(row=row_num, column=3).value = source_ip
    sheet.cell(row=row_num, column=4).value = dest_name
    sheet.cell(row=row_num, column=5).value = dest_ip
    sheet.cell(row=row_num, column=6).value = port
    sheet.cell(row=row_num, column=7).value = protocol
    sheet.cell(row=row_num, column=8).value = requestor
    sheet.cell(row=row_num, column=9).value = "ITSR Consolidation Test"


def _fake_tickets_split(source_ip, destination_ip, return_list=True):
    if return_list:
        return ["ITSR"]
    return f"{source_ip}->{destination_ip}"


class MultiSplitItsrConsolidationTests(unittest.TestCase):
    @patch("auto_tickets.views.multi_split.tickets_split", side_effect=_fake_tickets_split)
    def test_consolidates_cfg_and_propagates_requestor(self, _mock_split):
        sheet = _build_sheet()
        _set_itsr_row(sheet, 4, "src-a", "10.10.10.1", "dst-a", "10.20.20.1", "443", "TCP", "")
        _set_itsr_row(sheet, 5, "src-a", "10.10.10.1", "dst-b", "10.20.20.2", "443", "TCP", "S123456")

        data = _process_itsr_file(sheet)

        self.assertEqual(len(data["itsr_sip_dic"]), 1)
        self.assertEqual(data["itsr_sip_dic"][4], "10.10.10.1")
        self.assertEqual(data["itsr_dip_dic"][4], "10.20.20.1\n10.20.20.2")
        self.assertEqual(set(data["itsr_requestor_dic"].values()), {"S123456"})

    @patch("auto_tickets.views.multi_split.tickets_split", side_effect=_fake_tickets_split)
    def test_cfg_efg_or_linking_merges_into_single_group(self, _mock_split):
        sheet = _build_sheet()
        _set_itsr_row(sheet, 4, "src-a", "10.10.10.1", "dst-x", "10.20.20.1", "443", "TCP", "S123456")
        _set_itsr_row(sheet, 5, "src-a", "10.10.10.1", "dst-y", "10.20.20.2", "443", "TCP", "S123456")
        _set_itsr_row(sheet, 6, "src-b", "10.10.10.2", "dst-y", "10.20.20.2", "443", "TCP", "S123456")

        data = _process_itsr_file(sheet)

        self.assertEqual(len(data["itsr_sip_dic"]), 1)
        self.assertEqual(data["itsr_sip_dic"][4], "10.10.10.1\n10.10.10.2")
        self.assertEqual(data["itsr_dip_dic"][4], "10.20.20.1\n10.20.20.2")

    @patch("auto_tickets.views.multi_split.tickets_split", side_effect=_fake_tickets_split)
    def test_non_matching_rows_remain_separate(self, _mock_split):
        sheet = _build_sheet()
        _set_itsr_row(sheet, 4, "src-a", "10.10.10.1", "dst-a", "10.20.20.1", "443", "TCP", "S1")
        _set_itsr_row(sheet, 5, "src-b", "10.10.10.2", "dst-b", "10.20.20.2", "8443", "TCP", "S1")

        data = _process_itsr_file(sheet)

        self.assertEqual(len(data["itsr_sip_dic"]), 2)


if __name__ == "__main__":
    unittest.main()
