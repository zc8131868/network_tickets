import ipaddress
import math
import django
import os
import sys

# Add the parent directory to Python path so we can import network_tickets.settings
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'network_tickets.settings')
django.setup()
from auto_tickets.models import IPDB, IP_Application, PA_Service


# print(network)
# print(type(network))

# ip_mask = [{'ip_prefix': '10.244.0.0', 'mask': ' 255.255.248.0'}, 
#         {'ip_prefix': '2.1.1.2', 'mask': '255.255.255.255'}]

def get_location(ip_input):
    if not ip_input:
        return None
        
    try:
        ip_mask = [{'ip_prefix': i.ip, 'mask': i.mask} for i in IPDB.objects.all()]        

        network_list = []

        for i in ip_mask:
            network = ipaddress.ip_network(f"{i['ip_prefix']}/{i['mask'].replace(' ', '')}")
            network_list.append(network)

        ip = ipaddress.ip_address(ip_input)

        for subnet in network_list:
            if ip in subnet:
                location = IPDB.objects.get(ip=subnet.network_address).location
                return location
        
    except (ValueError, ipaddress.AddressValueError):
        return None 


def get_device(ip_input):
    if not ip_input:
        return None
        
    try:
        ip_mask = [{'ip_prefix': i.ip, 'mask': i.mask} for i in IPDB.objects.all()]        

        network_list = []

        for i in ip_mask:
            network = ipaddress.ip_network(f"{i['ip_prefix']}/{i['mask'].replace(' ', '')}")
            network_list.append(network)

        ip = ipaddress.ip_address(ip_input)

        for subnet in network_list:
            if ip in subnet:
                device = IPDB.objects.get(ip=subnet.network_address).device
                return device
        
    except (ValueError, ipaddress.AddressValueError):
        return None 


def tickets_split(source_ip, destination_ip):

    '''
    DMZ SW01: SN OAM, AliCloud,AliCloud-Mylink
    DMZ SW02: NewPrivateCloud, PrivateCloud, SN PCloud
    M09-CORE-SW01: PrivateCloud
    M09-EXT-CORE-SW1: PrivateCloud
    M09-INT-SW01: PrivateCloud
    M09-SB-SW01: South Base
    PA: South Base
    T01-DR-CORE-SW01: PrivateCloud-GNC
    '''

    source_location = get_location(source_ip)
    destination_location = get_location(destination_ip)
    
    source_device = get_device(source_ip)
    destination_device = get_device(destination_ip)

    # Check if we got valid locations
    if source_location is None or destination_location is None:
        return 'Unknown IP. Please report to IT, thank you.'
     #DMZ SW01
    if source_device == 'DMZ SW01' and destination_device == 'DMZ SW02':
        if source_location == 'SN OAM' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n'
        elif source_location == 'SN OAM' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n'
        elif source_location == 'SN OAM' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n'
        elif source_location == 'AliCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'
        elif source_location == 'AliCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'
        elif source_location == 'AliCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'
    
    elif source_device == 'DMZ SW01' and destination_device == 'M09-CORE-SW01':
        if source_location == 'SN OAM' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'AliCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'

    elif source_device == 'DMZ SW01' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'SN OAM' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'AliCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'

    elif source_device == 'DMZ SW01' and destination_device == 'M09-INT-SW01':
        if source_location == 'SN OAM' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'AliCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'

    elif source_device == 'DMZ SW01' and destination_device == 'M09-SB-SW01':
        if source_location == 'SN OAM' and destination_location == 'South Base':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'AliCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'
        # elif source_location == 'AliCloud-Mylink' and destination_location == 'South Base':
        #     return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'

    elif source_device == 'DMZ SW01' and destination_device == 'PA':
        if source_location == 'SN OAM' and destination_location == 'South Base':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'AliCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'
        # elif source_location == 'AliCloud-Mylink' and destination_location == 'South Base':
        #     return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'

    elif source_device == 'DMZ SW01' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'SN OAM' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'AliCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'
            
    elif source_device == 'DMZ SW02' and destination_device == 'M09-CORE-SW01':
        if source_location == 'SN PCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        elif source_location == 'NewPrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to NewPrivateCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    elif source_device == 'DMZ SW02' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'SN PCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        elif source_location == 'NewPrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to NewPrivateCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        
    elif source_device == 'DMZ SW02' and destination_device == 'M09-INT-SW01':
        if source_location == 'SN PCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        elif source_location == 'NewPrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to NewPrivateCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    elif source_device == 'DMZ SW02' and destination_device == 'M09-SB-SW01':
        if source_location == 'SN PCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'NewPrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to NewPrivateCloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'

    elif source_device == 'DMZ SW02' and destination_device == 'PA':
        if source_location == 'SN PCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'NewPrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to NewPrivateCloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'

    elif source_device == 'DMZ SW02' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'SN PCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'    
        elif source_location == 'NewPrivateCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to NewPrivateCloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    elif source_device == 'M09-CORE-SW01' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-CORE-SW01' and destination_device == 'M09-INT-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-CORE-SW01' and destination_device == 'M09-SB-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'M09-CORE-SW01' and destination_device == 'PA':
        if source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'M09-CORE-SW01' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'M09-INT-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'M09-SB-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'PA':
        if source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-INT-SW01' and destination_device == 'M09-SB-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'M09-INT-SW01' and destination_device == 'PA':
        if source_location == 'PrivateCloud' and destination_location == 'South Base':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'M09-INT-SW01' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-SB-SW01' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'PA' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    # Reversed device combinations
    elif source_device == 'DMZ SW02' and destination_device == 'DMZ SW01':
        if source_location == 'NewPrivateCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n'
        elif source_location == 'PrivateCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n'
        elif source_location == 'SN PCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n'
        elif source_location == 'NewPrivateCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'
        elif source_location == 'SN PCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)AliCloud'


    elif source_device == 'M09-CORE-SW01' and destination_device == 'DMZ SW01':
        if source_location == 'PrivateCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'

    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'DMZ SW01':
        if source_location == 'PrivateCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'

    elif source_device == 'M09-INT-SW01' and destination_device == 'DMZ SW01':
        if source_location == 'PrivateCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'PrivateCloud' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'

    elif source_device == 'M09-SB-SW01' and destination_device == 'DMZ SW01':
        if source_location == 'South Base' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'South Base' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'
        # elif source_location == 'South Base' and destination_location == 'AliCloud-Mylink':
        #     return f'{source_ip} belongs to South Base, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'

    elif source_device == 'PA' and destination_device == 'DMZ SW01':
        if source_location == 'South Base' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'South Base' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'
        # elif source_location == 'South Base' and destination_location == 'AliCloud-Mylink':
        #     return f'{source_ip} belongs to South Base, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)\n 4)AliCloud'

    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'DMZ SW01':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud-GNC' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR\n 4)AliCloud'
        elif source_location == 'PrivateCloud-GNC' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)AliCloud'

    elif source_device == 'M09-CORE-SW01' and destination_device == 'DMZ SW02':
        if source_location == 'PrivateCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to NewPrivateCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'DMZ SW02':
        if source_location == 'PrivateCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to NewPrivateCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        
    elif source_device == 'M09-INT-SW01' and destination_device == 'DMZ SW02':
        if source_location == 'PrivateCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
        elif source_location == 'PrivateCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to NewPrivateCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    elif source_device == 'M09-SB-SW01' and destination_device == 'DMZ SW02':
        if source_location == 'South Base' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'South Base' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to NewPrivateCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'

    elif source_device == 'PA' and destination_device == 'DMZ SW02':
        if source_location == 'South Base' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-SN \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
        elif source_location == 'South Base' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to NewPrivateCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'

    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'DMZ SW02':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN \n 3)ITSR'
        elif source_location == 'PrivateCloud-GNC' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'    
        elif source_location == 'PrivateCloud-GNC' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to NewPrivateCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'

    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'M09-CORE-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-INT-SW01' and destination_device == 'M09-CORE-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-SB-SW01' and destination_device == 'M09-CORE-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'PA' and destination_device == 'M09-CORE-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'M09-CORE-SW01':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-INT-SW01' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-SB-SW01' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'PA' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'M09-SB-SW01' and destination_device == 'M09-INT-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'PA' and destination_device == 'M09-INT-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR\n 3)South Base (IT will provide support)'
    
    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'M09-INT-SW01':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'M09-SB-SW01':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'South Base':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'PA':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'South Base':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to South Base. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    # Additional missing reversed combinations
    elif source_device == 'M09-SB-SW01' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    elif source_device == 'PA' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'South Base' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud \n 2)ITSR'
    
    # Self-to-self combinations (same device to same device)
    elif source_device == 'DMZ SW01' and destination_device == 'DMZ SW01':
        if source_location == 'SN OAM' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-SN'
        elif source_location == 'AliCloud' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)AliCloud'
        elif source_location == 'SN OAM' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)EOMS-SN \n 2)AliCloud'
        elif source_location == 'AliCloud' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-SN \n 2)AliCloud'
        elif source_location == 'SN OAM' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to SN OAM, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)EOMS-SN \n 2)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'SN OAM':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to SN OAM. Tickets contain: \n 1)EOMS-SN \n 2)AliCloud'
        elif source_location == 'AliCloud' and destination_location == 'AliCloud-Mylink':
            return f'{source_ip} belongs to AliCloud, {destination_ip} belongs to AliCloud-Mylink. Tickets contain: \n 1)AliCloud'
        elif source_location == 'AliCloud-Mylink' and destination_location == 'AliCloud':
            return f'{source_ip} belongs to AliCloud-Mylink, {destination_ip} belongs to AliCloud. Tickets contain: \n 1)AliCloud'
    
    elif source_device == 'DMZ SW02' and destination_device == 'DMZ SW02':
        if source_location == 'NewPrivateCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
        elif source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
        elif source_location == 'SN PCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-SN'
        elif source_location == 'NewPrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
        elif source_location == 'PrivateCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
        elif source_location == 'NewPrivateCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN'
        elif source_location == 'SN PCloud' and destination_location == 'NewPrivateCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN'
        elif source_location == 'PrivateCloud' and destination_location == 'SN PCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to SN PCloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN'
        elif source_location == 'SN PCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to SN PCloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud \n 2)EOMS-SN'
    
    elif source_device == 'M09-CORE-SW01' and destination_device == 'M09-CORE-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
    
    elif source_device == 'M09-EXT-CORE-SW1' and destination_device == 'M09-EXT-CORE-SW1':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
    
    elif source_device == 'M09-INT-SW01' and destination_device == 'M09-INT-SW01':
        if source_location == 'PrivateCloud' and destination_location == 'PrivateCloud':
            return f'{source_ip} belongs to Private Cloud, {destination_ip} belongs to Private Cloud. Tickets contain: \n 1)EOMS-Cloud'
    
    elif source_device == 'M09-SB-SW01' and destination_device == 'M09-SB-SW01':
        if source_location == 'South Base' and destination_location == 'South Base':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to South Base. Tickets contain: \n 1)South Base (IT will provide support)'
    
    elif source_device == 'PA' and destination_device == 'PA':
        if source_location == 'South Base' and destination_location == 'South Base':
            return f'{source_ip} belongs to South Base, {destination_ip} belongs to South Base. Tickets contain: \n 1)South Base (IT will provide support)'
    
    elif source_device == 'T01-DR-CORE-SW01' and destination_device == 'T01-DR-CORE-SW01':
        if source_location == 'PrivateCloud-GNC' and destination_location == 'PrivateCloud-GNC':
            return f'{source_ip} belongs to PrivateCloud-GNC, {destination_ip} belongs to PrivateCloud-GNC. Tickets contain: \n 1)EOMS-Cloud'
    
    else:
        return 'Unknown IP. Please report to IT, thank you.'




def generate_subnet(network, num_ip_addresses, existing_subnets=None):
    """
    Generate a subnet from a given network with enough capacity for specified IPs and check for overlaps.
    
    Args:
        network (str or ipaddress.IPv4Network/IPv6Network): The parent network
        num_ip_addresses (int): Number of IP addresses needed in the new subnet
        existing_subnets (list): List of existing subnets to check against
        
    Returns:
        ipaddress.IPv4Network/IPv6Network: The generated subnet if no overlap
        None: If no valid subnet can be generated or there's an overlap
        
    Raises:
        ValueError: If input parameters are invalid
    """
    # Validate number of IP addresses
    if not isinstance(num_ip_addresses, int) or num_ip_addresses <= 0:
        raise ValueError("Number of IP addresses must be a positive integer")
        
    # Normalize input network to ipaddress object
    if isinstance(network, str):
        network = ipaddress.ip_network(network, strict=False)
    elif not isinstance(network, (ipaddress.IPv4Network, ipaddress.IPv6Network)):
        raise ValueError("Invalid network format. Provide a string or ipaddress Network object.")
    
    # Calculate required prefix length
    # We need to account for network and broadcast addresses in IPv4 (hence +2)
    if isinstance(network, ipaddress.IPv4Network):
        required_hosts = num_ip_addresses + 2  # +2 for network and broadcast addresses
    else:  # IPv6 doesn't use broadcast addresses
        required_hosts = num_ip_addresses
    
    # Calculate minimum prefix length needed
    if required_hosts <= 1:
        prefix_length = network.max_prefixlen
    else:
        # Calculate the number of host bits needed
        host_bits = math.ceil(math.log2(required_hosts))
        prefix_length = network.max_prefixlen - host_bits
    
    # Validate prefix length
    if prefix_length <= network.prefixlen:
        raise ValueError(f"Not enough space in network {network} to accommodate {num_ip_addresses} IP addresses")
        
    if prefix_length > network.max_prefixlen:
        raise ValueError(f"Calculated prefix length ({prefix_length}) is invalid for this network type")
    
    # Initialize existing subnets list if not provided
    existing_subnets = existing_subnets or []
    
    # Convert existing subnets to ipaddress objects if they're strings
    normalized_existing = []
    for subnet in existing_subnets:
        if isinstance(subnet, str):
            normalized_existing.append(ipaddress.ip_network(subnet, strict=False))
        elif isinstance(subnet, (ipaddress.IPv4Network, ipaddress.IPv6Network)):
            normalized_existing.append(subnet)
        else:
            raise ValueError("Existing subnets must be strings or ipaddress Network objects")
    
    # Generate possible subnets and find the first available one
    for subnet in network.subnets(new_prefix=prefix_length):
        # Check if this subnet overlaps with any existing subnet
        overlap = any(subnet.overlaps(existing) for existing in normalized_existing)
        
        if not overlap:
            return subnet
            
    # If no available subnet found
    return None



import re
from netmiko import ConnectHandler


def get_nat_config(target_ip):
    firewall_PA = {'device_type': 'paloalto_panos',
                'host': '10.254.0.14',
                'username': 'ZhengCheng',
                'password': 'Cmhk@941'
                }
    net_connect = ConnectHandler(**firewall_PA)
    output = net_connect.send_command('show session all filter | match ' + target_ip)


    output_str = '''
                1380783      ssl            ACTIVE  FLOW  NS   172.19.1.193[64363]/Internal/6  (43.252.52.2[53865])
                1788626      ssl            ACTIVE  FLOW  NS   172.19.1.180[63100]/Internal/6  (43.252.52.2[40906])
                237626       wechat-base    ACTIVE  FLOW  NS   172.19.1.136[54176]/Internal/6  (43.252.52.2[51325])
                812034       ssl            ACTIVE  FLOW  NS   172.19.1.252[55376]/Internal/6  (43.252.52.2[9834])
                2388246      bittorrent     ACTIVE  FLOW  NS   172.19.1.224[51413]/Internal/17  (43.252.52.2[34251])
                1169254      bittorrent     ACTIVE  FLOW  NS   172.19.1.224[51413]/Internal/17  (43.252.52.2[8335])
                2249456      google-base    ACTIVE  FLOW  NS   172.19.1.237[52493]/Internal/6  (43.252.52.2[30334])
                vsys1                                          43.252.52.231[2123]/External  (43.252.52.231[2123])
                '''
    net_connect.disconnect()

    pa_res = re.findall(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\[(\d+)\][^(]*\((\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})\[(\d+)\]\)', output)
    
    return pa_res
    # for i in res:
    #     print(f'source_ip: {i[0]}')
    #     print(f'source_port: {i[1]}')
    #     print(f'destination_ip: {i[2]}')
    #     print(f'destination_port: {i[3]}')
    #     print('-' * 50)


#Auto Tickets PA
import re
from netmiko import ConnectHandler
from time import sleep
import openpyxl
import django
import os
import sys
import logging
from datetime import datetime

# Add the parent directory to Python path so we can import network_tickets.settings
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'network_tickets.settings')
django.setup()

# Setup logging
def setup_logging():
    """Setup comprehensive logging for Netmiko sessions"""
    # Create logs directory if it doesn't exist
    log_dir = '/it_network/network_tickets/logs'
    os.makedirs(log_dir, exist_ok=True)
    
    # Create timestamp for log files
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Setup main logger
    logger = logging.getLogger('netmiko_session')
    logger.setLevel(logging.DEBUG)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # File handler for session log
    session_log_file = f'{log_dir}/netmiko_session_{timestamp}.log'
    file_handler = logging.FileHandler(session_log_file)
    file_handler.setLevel(logging.DEBUG)
    
    # Console handler for real-time output
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Setup Netmiko debug logging
    netmiko_logger = logging.getLogger('netmiko')
    netmiko_logger.setLevel(logging.DEBUG)
    
    # Netmiko debug file handler
    debug_log_file = f'{log_dir}/netmiko_debug_{timestamp}.log'
    debug_handler = logging.FileHandler(debug_log_file)
    debug_handler.setLevel(logging.DEBUG)
    debug_handler.setFormatter(formatter)
    netmiko_logger.addHandler(debug_handler)
    
    logger.info(f"Logging initialized - Session log: {session_log_file}")
    logger.info(f"Debug log: {debug_log_file}")
    
    return logger



def auto_tickets_pa_tools(wb):
    # Initialize logging
    logger = setup_logging()
    logger.info("Starting auto_tickets_pa function")
    res_log = []
    res_log.append('Starting auto_tickets_pa function')
    # wb = openpyxl.load_workbook('/it_network/network_tickets/auto_tickets/sample.xlsx')
    # sheet = wb['Content']
    sheet = wb.active
    logger.info("Excel file loaded successfully")

    pattern = r'[ ,\n、]'

    start_row = 4
    end_row = sheet.max_row

    firewall_PA = {'device_type': 'paloalto_panos',
               'host': '10.254.0.14',
               'username': 'ZhengCheng',
               'password': 'Cmhk@941',
               'session_log': '/it_network/network_tickets/logs/netmiko_session.log',
               'session_log_file_mode': 'append',
               'global_delay_factor': 2,
               }
    
    logger.info(f"Connecting to firewall: {firewall_PA['host']}")
    try:
        net_connect = ConnectHandler(**firewall_PA)
        logger.info("Successfully connected to firewall")
        logger.info(f"Device prompt: {net_connect.find_prompt()}")
        res_log.append(f"Successfully connected to firewall")


    except Exception as e:
        logger.error(f"Failed to connect to firewall: {e}")
        res_log.append(f"Failed to connect to firewall: {e}")

        return res_log


    for row in range(start_row, end_row+1):
        
        ticket_number_dic = {}
        sip_dic = {}
        dip_dic = {}
        dport_dic = {}
        protocol_dic = {}

        ticket_number_list = []
        sip_list = []
        dip_list = []
        dport_list = []
        protocol_list = []
        print(f'row: {row}')
        ticket_number = sheet.cell(row=row, column=1).value
        sip = sheet.cell(row=row, column=3).value
        dip = sheet.cell(row=row, column=5).value
        dport = sheet.cell(row=row, column=7).value
        protocol = sheet.cell(row=row, column=6).value
        
        if ticket_number is not None:
            ticket_number_dic[row] = ticket_number
        else:
            raise Exception('ROW'+str(row)+' :Please provide the ticket number')
        

        if protocol is not None: 
            protocol_dic[row] = protocol
        else:
            raise Exception('ROW'+str(row)+' :Protocol is not defined')
        
        if sip is not None:
            for i in re.split(pattern, sip):
                try:
                    if re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', i):
                        sip_list.append(i)
                except:
                    raise Exception('ROW'+str(row)+' :Source IP is not defined')
            sip_dic[row] = sip_list
        else:
            raise Exception('ROW'+str(row)+' :Source IP is not defined')

        if dip is not None:
            for i in re.split(pattern, dip):
                try:
                    if re.search(r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}', i):
                        dip_list.append(i)
                except:
                    raise Exception('ROW'+str(row)+' :Destination IP is not defined')
            dip_dic[row] = dip_list
        else:
            raise Exception('ROW'+str(row)+' :Destination IP is not defined')
            
            
        if dport is not None:
            for i in re.split(pattern, str(dport)):
                if i != '':
                    dport_list.append(i)
            dport_dic[row] = dport_list
        else:
            raise Exception('ROW'+str(row)+' :Destination Port is not defined')


        sip_interface_dic = {}
        dip_interface_dic = {}

        for sip in sip_dic[row]:
            print(f'sip: {sip}')
            command = f'test routing fib-lookup virtual-router vr_vsys1 ip {sip}'
            logger.info(f"Executing command: {command}")
            check_sip_raw = net_connect.send_command(command)
            logger.debug(f'check_sip_raw response: {check_sip_raw}')
            res_log.append(f"Executing command: {command}")
            print(f'check_sip_raw: {check_sip_raw}')
            # print(f'check_sip_raw:{check_sip_raw}')

            # print('-' * 100)
            interface = re.search(r'interface\s(ethernet\d/\d+),',check_sip_raw).group(1) 

            # print(f'interface:{interface}')
            # print('-' * 100)
            logger.info(f"Source IP {sip} mapped to interface {interface}")
            res_log.append(f"Source IP {sip} mapped to interface {interface}")
            sip_interface_dic[sip] = interface


        for dip in dip_dic[row]:
            command = f'test routing fib-lookup virtual-router vr_vsys1 ip {dip}'
            logger.info(f"Executing command: {command}")
            check_dip_raw = net_connect.send_command(command)
            logger.debug(f'check_dip_raw response: {check_dip_raw}')
            res_log.append(f"Executing command: {command}")
            # print(check_dip_raw)
            interface = re.search(r'interface\s(ethernet\d/\d+),',check_dip_raw).group(1)    
            logger.info(f"Destination IP {dip} mapped to interface {interface}")
            res_log.append(f"Destination IP {dip} mapped to interface {interface}")
            dip_interface_dic[dip] = interface
        
    #get source ip and destination ip in reach row, if the interfaces of source ip and destination ip are different, then add to the set
        sip_ip_set = set()
        dip_ip_set = set()

        for sip, s_interface in sip_interface_dic.items():
            for dip, d_interface in dip_interface_dic.items():
                if s_interface != d_interface:
                    print(f'ticket number: {ticket_number_dic[row]}')
                    print(f'source interface: {s_interface}')
                    print(f'destination interface: {d_interface}')
                    print(f'source ip: {sip}')
                    print(f'destination ip: {dip}')
                    print(f'destination port: {dport_dic[row]}')
                    print('-' * 100)
                    res_log.append(f"ticket number: {ticket_number_dic[row]}")
                    res_log.append(f"source interface: {s_interface}")
                    res_log.append(f"destination interface: {d_interface}")
                    res_log.append(f"source ip: {sip}")
                    res_log.append(f"destination ip: {dip}")
                    res_log.append(f"destination port: {dport_dic[row]}")
                    sip_ip_set.add(sip)
                    dip_ip_set.add(dip)
                else:
                    print(f'{sip} and {dip} belong to the same zone. No action needed')
                    print('-' * 100)
                    res_log.append(f"{sip} and {dip} belong to the same zone. No action needed")

        #skip the current row if there is no action needed   
        if sip_ip_set ==set():
            continue

        print('-' * 100)
        print('Create source ip and destination ip in the firewall')
        res_log.append(f"Create source ip and destination ip in the firewall")
        print('-' * 100)
        
        print(f'source set: {sip_ip_set}')
        print(f'destination set: {dip_ip_set}')
        print('-' * 100)

        # print('enter config mode')
        logger.info("Entering configuration mode")
        res_log.append(f"Entering configuration mode")
        out = net_connect.config_mode()
        logger.debug(f"Config mode response: {out}")
        # print(out)
        # print('-' * 100)
        logger.info("Starting address creation in firewall")
        res_log.append(f"Starting address creation in firewall")
        #create source ip and destination ip in the firewall
        for sip in sip_ip_set:
            command = f'set address {sip} description generated-by-netcare ip-netmask {sip}'
            logger.info(f"Creating source address: {command}")
            res_log.append(f"Creating source address: {command}")
            try:
                res = net_connect.send_command(command)
                logger.info(f"Successfully created address {sip}")
                print(f"Created address {sip}: {res}")
                res_log.append(f"Successfully created address {sip}")
                
            except Exception as e:
                logger.error(f"Failed to create address {sip}: {e}")
                print(f"Failed to create address {sip}: {e}")
                res_log.append(f"Failed to create address {sip}: {e}")
                
        # Create destination addresses  
        for dip in dip_ip_set:
            command = f'set address {dip} description generated-by-netcare ip-netmask {dip}'
            logger.info(f"Creating destination address: {command}")
            res_log.append(f"Creating destination address: {command}")
            try:
                res = net_connect.send_command(command)
                logger.info(f"Successfully created address {dip}")
                print(f"Created address {dip}: {res}")
                res_log.append(f"Successfully created address {dip}")
            
            except Exception as e:
                logger.error(f"Failed to create address {dip}: {e}")
                print(f"Failed to create address {dip}: {e}")
                res_log.append(f"Failed to create address {dip}: {e}")
                
        print('-' * 100)
        res_log.append(f"Create source address-group and destination address-group in the firewall")
        #create source address-group and destination address-group in the firewall
        print('Create source address-group and destination address-group in the firewall')
        sleep(5)
        # print("Back to operational mode. Prompt:", net_connect.find_prompt())

        # print(type(ticket_number_dic[row]))
        print("Back to operational mode. Prompt:", net_connect.find_prompt())

        command = f'edit address-group {ticket_number_dic[row]}-src-row{row}'
        logger.info(f"Creating source address-group: {command}")
        res_log.append(f"Creating source address-group: {command}")
        res = net_connect.send_command(command)
        logger.info(f"Successfully created source address-group: {res}")
        res_log.append(f"Successfully created source address-group: {res}")
        print('-' * 100)
        for sip in sip_ip_set:
            command = f'set static {sip}'
            res = net_connect.send_command(command)
            logger.info(f"Successfully added static address: {res}")
            res_log.append(f"Successfully added static address: {res}")
            print(f'Added static address: {sip}')
        command = f'set description generated-by-netcare'
        res = net_connect.send_command(command)
        logger.info(f"Successfully set description: {res}")
        res_log.append(f"Successfully set description: {res}")
        net_connect.send_command('exit')
        print('-' * 100)    
        command = f'edit address-group {ticket_number_dic[row]}-dst-row{row}'
        logger.info(f"Creating destination address-group: {command}")
        res_log.append(f"Creating destination address-group: {command}")
        res = net_connect.send_command(command)
        logger.info(f"Successfully created destination address-group: {res}")
        print('-' * 100)
        sleep(2)
        for dip in dip_ip_set:
            command = f'set static {dip}'
            res_log.append(f"Adding static address: {command}")
            res = net_connect.send_command(command)
            logger.info(f"Successfully added static address: {res}")
            print(f'Added static address: {dip}')
            res_log.append(f"Successfully added static address: {res}")

        command = f'set description generated-by-netcare'
        res = net_connect.send_command(command)
        logger.info(f"Successfully set description: {res}")
        res_log.append(f"Successfully set description: {res}")

        # # #create service and security rule in the firewall
        print('create service and security rule in the firewall')
        res_log.append(f"create service and security rule in the firewall")

        sleep(1)
        net_connect.send_command('exit')
        # print('create service in the firewall')
        if protocol_dic[row] == 'tcp' or protocol_dic[row] == 'TCP':
            for dport in dport_dic[row]:
                print(f'dport: {dport}')
                print(f'dport_dic[row]: {dport_dic[row]}')
                if dport != 'icmp' and dport != 'ICMP':
                    command = f'set service tcp-{dport}-netcare description generated-by-netcare protocol tcp port {dport}'
                    logger.info(f"Creating service: {command}")
                    res_log.append(f"Creating service: {command}")
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully created service: {res}")
                    res_log.append(f"Successfully created service: {res}")

                    command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-tcp' + dport + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + ' from any to any service ' + 'tcp-' + dport + '-netcare',
                    'set description generated-by-netcare','set application any','set action allow','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")

                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")

                    # print("Back to operational mode. Prompt:", net_connect.find_prompt())

                            
                else:
                    command_list = ['edit rulebase security rules ' + 'icmp' + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + ' from any to any service application-default application icmp ',
                    'set description generated-by-netcare','set action allow','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")

        elif protocol_dic[row] == 'udp' or protocol_dic[row] == 'UDP':
            for dport in dport_dic[row]:
                if dport != 'icmp' and dport != 'ICMP':
                    command = f'set service udp-{dport}-netcare description generated-by-netcare protocol udp port {dport}'
                    logger.info(f"Creating service: {command}")
                    res_log.append(f"Creating service: {command}")
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully created service: {res}")
                    res_log.append(f"Successfully created service: {res}")
                    
                    command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-udp' + dport + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + ' from any to any service ' + 'udp-' + dport + '-netcare',
                    'set description generated-by-netcare','set application any','set action allow','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                    
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")
                    res_log.append(f"Successfully created security rule: {res}")
                else:
                    command_list = ['edit rulebase security rules ' + 'icmp' + '-' + 'netcare', 
                    'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + ' from any to any service application-default application icmp ',
                    'set description generated-by-netcare','set action allow','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")
                    res_log.append(f"Successfully created security rule: {res}")
        elif protocol_dic[row] == 'icmp' or protocol_dic[row] == 'ICMP':
            command_list = ['edit rulebase security rules ' + 'icmp' + '-' + 'netcare', 
            'set source ' + ticket_number_dic[row] + '-src-row' + str(row) + ' destination ' + ticket_number_dic[row] + '-dst-row' + str(row) + ' from any to any service application-default application icmp ',
            'set description generated-by-netcare','set action allow','exit', 'exit', 'exit']
            res_log.append(f"Creating security rule: {command_list}")

            res = net_connect.send_config_set(command_list, exit_config_mode=False)
            logger.info(f"Successfully created security rule: {res}")
            res_log.append(f"Successfully created security rule: {res}")
        else:
            raise Exception('Pls check the protocol and port in the excel file.')
            res_log.append(f"Pls check the protocol and port in the excel file.")

        
        print('I am here')
        res_log.append(f"Exiting configuration mode")
        logger.info("Exiting configuration mode")
        # print("Back to operational mode. Prompt:", net_connect.find_prompt())
        net_connect.exit_config_mode()   
 
       

    net_connect.config_mode()
    # logger.info("Back to operational mode. Prompt:", net_connect.find_prompt())

    commit_command = 'commit description commit-by-netcare'
    commit_output = net_connect.send_command_timing(commit_command,read_timeout=0)
    print(f"Successfully committed: {commit_output}")
    logger.info(f"Successfully committed: {commit_output}")
    res_log.append(f"{commit_output}") 
    logger.info("Disconnecting from firewall")
    net_connect.disconnect()
    res_log.append(f"Disconnecting from firewall")
    logger.info("Successfully disconnected from firewall")
    logger.info("auto_tickets_pa function completed")
    output ='''
        {4: ['111-111'], 5: ['222-222']}
        {4: ['10.250.122.60', '10.250.122.5', '10.250.122.6', '10.250.122.9', '10.250.122.10', '10.250.122.11', '10.250.122.13', '10.250.122.18'], 5: ['10.0.57.21', '10.0.57.31']}
        {4: ['10.0.28.138'], 5: ['203.142.100.51', '203.142.100.52', '203.142.100.103', '203.142.100.104']}
        {4: ['3306', 'ICMP'], 5: ['53', '54']}
   '''
    return res_log




                
                    
            
   




