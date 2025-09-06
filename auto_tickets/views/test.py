import re
from netmiko import ConnectHandler
from time import sleep
import openpyxl
import django
import os
import sys
import logging
from datetime import datetime
from pprint import pprint

# Add the parent directory to Python path so we can import network_tickets.settings
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'network_tickets.settings')
django.setup()

# Setup logging
def setup_logging():
    """Setup comprehensive logging for Netmiko sessions"""
    # Create logs directory if it doesn't exist
    log_dir = '/it_network/network_tickets/logs'
    os.makedirs(log_dir, exist_ok=True)
    
    # Create timestamp for log files
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Setup main logger
    logger = logging.getLogger('netmiko_session')
    logger.setLevel(logging.DEBUG)
    
    # Clear any existing handlers
    logger.handlers.clear()
    
    # File handler for session log
    session_log_file = f'{log_dir}/netmiko_session_{timestamp}.log'
    file_handler = logging.FileHandler(session_log_file)
    file_handler.setLevel(logging.DEBUG)
    
    # Console handler for real-time output
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    
    # Formatter
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)
    
    # Add handlers
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    
    # Setup Netmiko debug logging
    netmiko_logger = logging.getLogger('netmiko')
    netmiko_logger.setLevel(logging.DEBUG)
    
    # Netmiko debug file handler
    debug_log_file = f'{log_dir}/netmiko_debug_{timestamp}.log'
    debug_handler = logging.FileHandler(debug_log_file)
    debug_handler.setLevel(logging.DEBUG)
    debug_handler.setFormatter(formatter)
    netmiko_logger.addHandler(debug_handler)
    
    logger.info(f"Logging initialized - Session log: {session_log_file}")
    logger.info(f"Debug log: {debug_log_file}")
    
    return logger



def auto_tickets_pa():
    # Initialize logging
    logger = setup_logging()
    logger.info("Starting auto_tickets_pa function")
    res_log = []
    res_log.append('Starting auto_tickets_pa function')
    wb = openpyxl.load_workbook('/it_network/network_tickets/auto_tickets/sample.xlsx')
    sheet = wb['Content']
    logger.info("Excel file loaded successfully")

    pattern = r'[ ,\n]'

    start_row = 4
    end_row = sheet.max_row

    firewall_PA = {'device_type': 'paloalto_panos',
               'host': '10.254.0.14',
               'username': 'ZhengCheng',
               'password': 'Cmhk@941',
               'session_log': '/it_network/network_tickets/logs/netmiko_session.log',
               'session_log_file_mode': 'append'
               }
    
    logger.info(f"Connecting to firewall: {firewall_PA['host']}")
    try:
        net_connect = ConnectHandler(**firewall_PA)
        logger.info("Successfully connected to firewall")
        logger.info(f"Device prompt: {net_connect.find_prompt()}")
        res_log.append(f"Successfully connected to firewall")
    except Exception as e:
        logger.error(f"Failed to connect to firewall: {e}")
        res_log.append(f"Failed to connect to firewall: {e}")
        return res_log


    for row in range(start_row, end_row+1):
        
        ticket_number_dic = {}
        sip_dic = {}
        dip_dic = {}
        dport_dic = {}
        protocol_dic = {}

        ticket_number_list = []
        sip_list = []
        dip_list = []
        dport_list = []
        protocol_list = []
        print(f'row: {row}')
        ticket_number = sheet.cell(row=row, column=1).value
        sip = sheet.cell(row=row, column=3).value
        dip = sheet.cell(row=row, column=5).value
        dport = sheet.cell(row=row, column=7).value
        protocol = sheet.cell(row=row, column=6).value
        
        if ticket_number is not None:
            ticket_number_dic[row] = ticket_number
        else:
            raise Exception('ROW'+str(row)+' :Please provide the ticket number')
        

        if protocol is not None: 
            protocol_dic[row] = protocol
        else:
            raise Exception('ROW'+str(row)+' :Protocol is not defined')
        
        if sip is not None:
            for i in re.split(pattern, sip):
                if i != '':
                    sip_list.append(i)
            sip_dic[row] = sip_list
        else:
            raise Exception('ROW'+str(row)+' :Source IP is not defined')

        if dip is not None:
            for i in re.split(pattern, dip):
                if i != '':
                    dip_list.append(i)
            dip_dic[row] = dip_list
        else:
            raise Exception('ROW'+str(row)+' :Destination IP is not defined')
            
            
        if dport is not None:
            for i in re.split(pattern, dport):
                if i != '':
                    dport_list.append(i)
            dport_dic[row] = dport_list
        else:
            raise Exception('ROW'+str(row)+' :Destination Port is not defined')


        sip_interface_dic = {}
        dip_interface_dic = {}

        for sip in sip_dic[row]:
            print(f'sip: {sip}')
            command = f'test routing fib-lookup virtual-router vr_vsys1 ip {sip}'
            logger.info(f"Executing command: {command}")
            check_sip_raw = net_connect.send_command(command)
            logger.debug(f'check_sip_raw response: {check_sip_raw}')
            res_log.append(f"Executing command: {command}")
            print(f'check_sip_raw: {check_sip_raw}')
            # print(f'check_sip_raw:{check_sip_raw}')

            # print('-' * 100)
            interface = re.search(r'interface\s(ethernet\d/\d+),',check_sip_raw).group(1) 

            # print(f'interface:{interface}')
            # print('-' * 100)
            logger.info(f"Source IP {sip} mapped to interface {interface}")
            res_log.append(f"Source IP {sip} mapped to interface {interface}")
            sip_interface_dic[sip] = interface


        for dip in dip_dic[row]:
            command = f'test routing fib-lookup virtual-router vr_vsys1 ip {dip}'
            logger.info(f"Executing command: {command}")
            check_dip_raw = net_connect.send_command(command)
            logger.debug(f'check_dip_raw response: {check_dip_raw}')
            res_log.append(f"Executing command: {command}")
            # print(check_dip_raw)
            interface = re.search(r'interface\s(ethernet\d/\d+),',check_dip_raw).group(1)    
            logger.info(f"Destination IP {dip} mapped to interface {interface}")
            res_log.append(f"Destination IP {dip} mapped to interface {interface}")
            dip_interface_dic[dip] = interface
        
    #get source ip and destination ip in reach row, if the interfaces of source ip and destination ip are different, then add to the set
        sip_ip_set = set()
        dip_ip_set = set()

        for sip, s_interface in sip_interface_dic.items():
            for dip, d_interface in dip_interface_dic.items():
                if s_interface != d_interface:
                    print(f'ticket number: {ticket_number_dic[row]}')
                    print(f'source interface: {s_interface}')
                    print(f'destination interface: {d_interface}')
                    print(f'source ip: {sip}')
                    print(f'destination ip: {dip}')
                    print(f'destination port: {dport_dic[row]}')
                    print('-' * 100)
                    res_log.append(f"ticket number: {ticket_number_dic[row]}")
                    res_log.append(f"source interface: {s_interface}")
                    res_log.append(f"destination interface: {d_interface}")
                    res_log.append(f"source ip: {sip}")
                    res_log.append(f"destination ip: {dip}")
                    res_log.append(f"destination port: {dport_dic[row]}")
                    res_log.append('-' * 100)
                    sip_ip_set.add(sip)
                    dip_ip_set.add(dip)
                else:
                    print(f'{sip} and {dip} belong to the same zone. No action needed')
                    print('-' * 100)
                    res_log.append(f"{sip} and {dip} belong to the same zone. No action needed")
                    res_log.append('-' * 100)

        #skip the current row if there is no action needed   
        if sip_ip_set ==set():
            continue

        print('-' * 100)
        print('Create source ip and destination ip in the firewall')
        res_log.append(f"Create source ip and destination ip in the firewall")
        print('-' * 100)
        
        print(f'source set: {sip_ip_set}')
        print(f'destination set: {dip_ip_set}')
        print('-' * 100)

        # print('enter config mode')
        logger.info("Entering configuration mode")
        res_log.append(f"Entering configuration mode")
        out = net_connect.config_mode()
        logger.debug(f"Config mode response: {out}")
        # print(out)
        # print('-' * 100)
        logger.info("Starting address creation in firewall")
        res_log.append(f"Starting address creation in firewall")
        #create source ip and destination ip in the firewall
        for sip in sip_ip_set:
            command = f'set address {sip} description generated-by-netcare ip-netmask {sip}'
            logger.info(f"Creating source address: {command}")
            res_log.append(f"Creating source address: {command}")
            try:
                res = net_connect.send_command(command)
                logger.info(f"Successfully created address {sip}")
                print(f"Created address {sip}: {res}")
                res_log.append(f"Successfully created address {sip}")
            except Exception as e:
                logger.error(f"Failed to create address {sip}: {e}")
                print(f"Failed to create address {sip}: {e}")
                res_log.append(f"Failed to create address {sip}: {e}")
        # Create destination addresses  
        for dip in dip_ip_set:
            command = f'set address {dip} description generated-by-netcare ip-netmask {dip}'
            logger.info(f"Creating destination address: {command}")
            res_log.append(f"Creating destination address: {command}")
            try:
                res = net_connect.send_command(command)
                logger.info(f"Successfully created address {dip}")
                print(f"Created address {dip}: {res}")
                res_log.append(f"Successfully created address {dip}")
            except Exception as e:
                logger.error(f"Failed to create address {dip}: {e}")
                print(f"Failed to create address {dip}: {e}")
                res_log.append(f"Failed to create address {dip}: {e}")
        
        print('-' * 100)
        res_log.append(f"Create source address-group and destination address-group in the firewall")
        #create source address-group and destination address-group in the firewall
        print('Create source address-group and destination address-group in the firewall')
        sleep(5)
        # print("Back to operational mode. Prompt:", net_connect.find_prompt())

        # print(type(ticket_number_dic[row]))
        print("Back to operational mode. Prompt:", net_connect.find_prompt())

        command = f'edit address-group {ticket_number_dic[row]}-source'
        logger.info(f"Creating source address-group: {command}")
        res_log.append(f"Creating source address-group: {command}")
        res = net_connect.send_command(command)
        logger.info(f"Successfully created source address-group: {res}")
        res_log.append(f"Successfully created source address-group: {res}")
        print('-' * 100)

        for sip in sip_ip_set:
            command = f'set static {sip}'
            res = net_connect.send_command(command)
            logger.info(f"Successfully added static address: {res}")
            res_log.append(f"Successfully added static address: {res}")
            print(f'Added static address: {sip}')
        command = f'set description generated-by-netcare'
        res = net_connect.send_command(command)
        logger.info(f"Successfully set description: {res}")
        res_log.append(f"Successfully set description: {res}")
        net_connect.send_command('exit')
        print('-' * 100)
        
        command = f'edit address-group {ticket_number_dic[row]}-destination'
        logger.info(f"Creating destination address-group: {command}")
        res_log.append(f"Creating destination address-group: {command}")
        res = net_connect.send_command(command)
        logger.info(f"Successfully created destination address-group: {res}")
        print('-' * 100)
        sleep(2)
        for dip in dip_ip_set:
            command = f'set static {dip}'
            res_log.append(f"Adding static address: {command}")
            res = net_connect.send_command(command)
            logger.info(f"Successfully added static address: {res}")
            print(f'Added static address: {dip}')
            res_log.append(f"Successfully added static address: {res}")
        command = f'set description generated-by-netcare'
        res = net_connect.send_command(command)
        logger.info(f"Successfully set description: {res}")
        res_log.append(f"Successfully set description: {res}")

        print('-' * 100)
        # # #create service and security rule in the firewall
        print('create service and security rule in the firewall')
        res_log.append(f"create service and security rule in the firewall")
        sleep(5)
        net_connect.send_command('exit')
        # print('create service in the firewall')
        if protocol_dic[row] == 'tcp' or protocol_dic[row] == 'TCP':
            for dport in dport_dic[row]:
                print(f'dport: {dport}')
                print(f'dport_dic[row]: {dport_dic[row]}')
                if dport != 'icmp' and dport != 'ICMP':
                    command = f'set service tcp-{dport}-netcare description generated-by-netcare protocol tcp port {dport}'
                    logger.info(f"Creating service: {command}")
                    res_log.append(f"Creating service: {command}")
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully created service: {res}")

                    command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-source' + ' destination ' + ticket_number_dic[row] + '-destination from any to any service ' + 'tcp-' + dport + '-netcare',
                    'set description generated-by-netcare','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")

                    # print("Back to operational mode. Prompt:", net_connect.find_prompt())

                            
                else:
                    command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-source' + ' destination ' + ticket_number_dic[row] + '-destination from any to any application icmp ',
                    'set description generated-by-netcare','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")

        elif protocol_dic[row] == 'udp' or protocol_dic[row] == 'UDP':
            for dport in dport_dic[row]:
                if dport != 'icmp' and dport != 'ICMP':
                    command = f'set service udp-{dport}-netcare description generated-by-netcare protocol udp port {dport}'
                    logger.info(f"Creating service: {command}")
                    res_log.append(f"Creating service: {command}")
                    res = net_connect.send_command(command)
                    logger.info(f"Successfully created service: {res}")
                    
                    command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-source' + ' destination ' + ticket_number_dic[row] + '-destination from any to any service ' + 'udp-' + dport + '-netcare',
                    'set description generated-by-netcare','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")

                else:
                    command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-netcare', 
                    'set source ' + ticket_number_dic[row] + '-source' + ' destination ' + ticket_number_dic[row] + '-destination from any to any application icmp ',
                    'set description generated-by-netcare','exit', 'exit', 'exit']
                    res_log.append(f"Creating security rule: {command_list}")
                    res = net_connect.send_config_set(command_list, exit_config_mode=False)
                    logger.info(f"Successfully created security rule: {res}")
                            
        elif protocol_dic[row] == 'icmp' or protocol_dic[row] == 'ICMP':
            command_list = ['edit rulebase security rules ' + ticket_number_dic[row] + '-netcare', 
            'set source ' + ticket_number_dic[row] + '-source' + ' destination ' + ticket_number_dic[row] + '-destination from any to any application icmp ',
            'set description generated-by-netcare','exit', 'exit', 'exit']
            res_log.append(f"Creating security rule: {command_list}")
            res = net_connect.send_config_set(command_list, exit_config_mode=False)
            logger.info(f"Successfully created security rule: {res}")
            
        else:
            raise Exception('Pls check the protocol and port in the excel file.')

        
        print('I am here')
        res_log.append(f"Exiting configuration mode")
        logger.info("Exiting configuration mode")
        # print("Back to operational mode. Prompt:", net_connect.find_prompt())
        net_connect.exit_config_mode()   
        # logger.info("Successfully exited configuration mode")
        print("Back to operational mode. Prompt:", net_connect.find_prompt())
        

    logger.info("Disconnecting from firewall")
    net_connect.disconnect()
    logger.info("Successfully disconnected from firewall")
    logger.info("auto_tickets_pa function completed")
    output ='''
        {4: ['111-111'], 5: ['222-222']}
        {4: ['10.250.122.60', '10.250.122.5', '10.250.122.6', '10.250.122.9', '10.250.122.10', '10.250.122.11', '10.250.122.13', '10.250.122.18'], 5: ['10.0.57.21', '10.0.57.31']}
        {4: ['10.0.28.138'], 5: ['203.142.100.51', '203.142.100.52', '203.142.100.103', '203.142.100.104']}
        {4: ['3306', 'ICMP'], 5: ['53', '54']}
   '''
    return res_log
        
log = auto_tickets_pa()
pprint(f'log: {log}')
                
                    
            
   




