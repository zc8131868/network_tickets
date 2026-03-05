#!/usr/bin/env python3
"""
Fill Username (P) and Password (Q) in CMHK IT网络设备信息 Excel for rows where
Login Method is "ssh" only. All other values are ignored.

For each target IP, every (username, password) pair from the login sheet is tried
until one succeeds; if none work, P and Q are left blank. Uses a thread pool to
test multiple IPs in parallel.

Usage:
  python ssh_fill_cmhk_excel.py [path_to_excel]
"""

import argparse
import logging
import os
import re
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import paramiko

# Accept SSH-2.99 (common on network devices) as compatible with 2.0
_paramiko_transport = paramiko.transport.Transport

# Re-implement _check_banner to allow version "2.99" in addition to "1.99" and "2.0"
def _check_banner_accept_299(self):
    from paramiko.common import DEBUG, INFO
    for i in range(100):
        timeout = self.banner_timeout if i == 0 else 2
        try:
            buf = self.packetizer.readline(timeout)
        except paramiko.ssh_exception.ProxyCommandFailure:
            raise
        except Exception as e:
            raise paramiko.ssh_exception.SSHException(
                "Error reading SSH protocol banner" + str(e)
            )
        if buf[:4] == "SSH-":
            break
        self._log(DEBUG, "Banner: " + buf)
    if buf[:4] != "SSH-":
        raise paramiko.ssh_exception.SSHException('Indecipherable protocol version "' + buf + '"')
    self.remote_version = buf
    self._log(DEBUG, "Remote version/idstring: {}".format(buf))
    i = buf.find(" ")
    if i >= 0:
        buf = buf[:i]
    segs = buf.split("-", 2)
    if len(segs) < 3:
        raise paramiko.ssh_exception.SSHException("Invalid SSH banner")
    version = segs[1]
    client = segs[2]
    if version not in ("1.99", "2.0", "2.99"):
        msg = "Incompatible version ({} instead of 2.0)"
        raise paramiko.ssh_exception.IncompatiblePeer(msg.format(version))
    msg = "Connected (version {}, client {})".format(version, client)
    self._log(INFO, msg)

_paramiko_transport._check_banner = _check_banner_accept_299

# Reduce paramiko log noise (avoid "Exception (client): ..." tracebacks on failed connects)
logging.getLogger("paramiko").setLevel(logging.WARNING)

SSH_TIMEOUT = 10
SSH_PORT = 22
DEFAULT_MAX_WORKERS = 10


def find_excel_path(given_path: str) -> str:
    """Resolve Excel file path; support Chinese filename."""
    if given_path and os.path.isfile(given_path):
        return given_path
    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    default_name = "CMHK IT网络设备信息-v1.xlsx"
    for base in (root, os.getcwd()):
        candidate = os.path.join(base, default_name)
        if os.path.isfile(candidate):
            return candidate
    for f in os.listdir(root):
        if "CMHK" in f and f.endswith(".xlsx"):
            return os.path.join(root, f)
    for f in os.listdir(os.getcwd()):
        if "CMHK" in f and f.endswith(".xlsx"):
            return os.path.join(os.getcwd(), f)
    return given_path or default_name


def get_login_credentials(ws_login) -> list:
    """Read (username, password) pairs from the 'login' sheet."""
    creds = []
    for r in range(2, ws_login.max_row + 1):
        u = ws_login.cell(r, 1).value
        p = ws_login.cell(r, 2).value
        if u is not None and str(u).strip():
            creds.append((str(u).strip(), str(p).strip() if p is not None else ""))
    return creds


def is_valid_ip(addr) -> bool:
    """Check if string looks like an IPv4 address or hostname."""
    if not addr:
        return False
    s = str(addr).strip()
    if not s or len(s) > 253:
        return False
    if re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", s):
        return True
    if re.match(r"^[a-zA-Z0-9.-]+$", s):
        return True
    return False


def try_ssh(host: str, username: str, password: str, port: int = SSH_PORT, timeout: int = SSH_TIMEOUT) -> bool:
    """Attempt SSH login; return True if successful."""
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    try:
        client.connect(
            hostname=host,
            port=port,
            username=username,
            password=password,
            timeout=timeout,
            allow_agent=False,
            look_for_keys=False,
        )
        return True
    except Exception:
        return False
    finally:
        try:
            client.close()
        except Exception:
            pass


def try_host_credentials(row_idx: int, host: str, creds: list, timeout: int = SSH_TIMEOUT) -> tuple:
    """
    Try each (username, password) pair for one host until one succeeds.
    Returns (row_idx, username, password) on success, or (row_idx, None, None) if no credential works.
    """
    for username, password in creds:
        if try_ssh(host, username, password, timeout=timeout):
            return (row_idx, username, password)
    return (row_idx, None, None)


def main():
    parser = argparse.ArgumentParser(description="Fill SSH username/password in CMHK device Excel (ssh rows only).")
    parser.add_argument("excel_path", nargs="?", default="", help="Path to CMHK IT网络设备信息-v1.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Do not save the workbook")
    parser.add_argument("--timeout", type=int, default=SSH_TIMEOUT, help="SSH connection timeout (seconds)")
    parser.add_argument("--workers", type=int, default=DEFAULT_MAX_WORKERS, help="Max concurrent SSH attempts (default %s)" % DEFAULT_MAX_WORKERS)
    args = parser.parse_args()

    path = find_excel_path(args.excel_path)
    if not os.path.isfile(path):
        print(f"Error: File not found: {path}", file=sys.stderr)
        sys.exit(1)

    print(f"Loading: {path}")
    wb = openpyxl.load_workbook(path)
    if "总表设备" not in wb.sheetnames:
        print('Error: Sheet "总表设备" not found.', file=sys.stderr)
        sys.exit(1)
    if "login" not in wb.sheetnames:
        print('Error: Sheet "login" not found.', file=sys.stderr)
        sys.exit(1)

    ws_main = wb["总表设备"]
    ws_login = wb["login"]

    headers = [ws_main.cell(1, c).value for c in range(1, ws_main.max_column + 1)]
    try:
        col_login_method = headers.index("Login Method") + 1
        col_manage_addr = headers.index("管理地址") + 1
        col_username = headers.index("Username") + 1
        col_password = headers.index("Password") + 1
    except ValueError as e:
        print(f"Error: Required column not found in 总表设备: {e}", file=sys.stderr)
        sys.exit(1)

    creds = get_login_credentials(ws_login)
    if not creds:
        print("Warning: No credentials found in 'login' sheet.")
    else:
        print(f"Loaded {len(creds)} credential(s) from 'login' sheet.")

    # Only process rows where Login Method is exactly "ssh". Ignore all other values (https, etc.).
    ssh_rows = []
    for r in range(2, ws_main.max_row + 1):
        lm = ws_main.cell(r, col_login_method).value
        if lm is None:
            continue
        if str(lm).strip().lower() != "ssh":
            continue  # ignore: not ssh
        addr = ws_main.cell(r, col_manage_addr).value
        if not is_valid_ip(addr):
            print(f"Row {r}: Skip (invalid 管理地址: {addr!r})")
            continue
        host = str(addr).strip()
        ssh_rows.append((r, host))

    print(f"Found {len(ssh_rows)} row(s) with Login Method = 'ssh' (all other rows ignored).")
    print(f"Using {args.workers} worker(s), testing all credential pairs per host until one works or none left.")

    # Run SSH attempts in parallel; each task tries all creds for one host until success or exhausted
    results = []
    with ThreadPoolExecutor(max_workers=args.workers) as executor:
        futures = {
            executor.submit(try_host_credentials, row_idx, host, creds, args.timeout): (row_idx, host)
            for row_idx, host in ssh_rows
        }
        for future in as_completed(futures):
            row_idx, username, password = future.result()
            row_idx_host = futures[future]
            results.append((row_idx, row_idx_host[1], username, password))

    # Apply results in row order and print
    results.sort(key=lambda x: x[0])
    for row_idx, host, username, password in results:
        if username is not None and password is not None:
            ws_main.cell(row_idx, col_username, username)
            ws_main.cell(row_idx, col_password, password)
            print(f"Row {row_idx} ({host}): OK -> {username}")
        else:
            ws_main.cell(row_idx, col_username, None)
            ws_main.cell(row_idx, col_password, None)
            print(f"Row {row_idx} ({host}): Failed (no credential worked, P/Q left blank)")

    if not args.dry_run and ssh_rows:
        wb.save(path)
        print(f"Saved: {path}")
    elif args.dry_run:
        print("Dry run: not saved.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
