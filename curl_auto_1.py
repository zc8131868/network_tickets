#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
域名拨测脚本 - 自动化版本
每1小时执行一次，每12次汇总输出Excel并计算平均值
"""

import subprocess
import re
import sys
import os
import time
import signal
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("正在安装 openpyxl 库...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl


# ============================================
# 配置区域
# ============================================
INTERVAL_MINUTES = 60  # 执行间隔（分钟）
BATCH_SIZE = 12  # 每多少次汇总一次
OUTPUT_DIR = "results"  # 输出目录

# ============================================
# 域名列表配置（在此处添加或修改域名）
# ============================================
DOMAINS = [
    "https://www.hk.chinamobile.com",
    "https://mymall.hk.chinamobile.com",
    "https://marketplace.hk.chinamobile.com",
    "https://omniapi.hk.chinamobile.com",
]


# 全局变量用于优雅退出
running = True


def signal_handler(signum, frame):
    """处理中断信号"""
    global running
    print("\n\n收到停止信号，正在退出...")
    running = False


def curl_test(url, user_agent="mylink/11", timeout=30):
    """执行curl拨测"""
    curl_cmd = [
        'curl',
        '-A', user_agent,
        '-o', '/dev/null',
        '-s',
        '-w', 'HTTP_CODE: %{http_code}\nDNS: %{time_namelookup}s\nTCP: %{time_connect}s\nSSL: %{time_appconnect}s\nTTFB: %{time_starttransfer}s\nTOTAL: %{time_total}s\n',
        '--max-time', str(timeout),
        url
    ]
    
    try:
        result = subprocess.run(curl_cmd, capture_output=True, text=True, timeout=timeout+5)
        output = result.stdout
        
        data = {
            'url': url,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'http_code': '',
            'dns': '',
            'tcp': '',
            'ssl': '',
            'ttfb': '',
            'total': ''
        }
        
        patterns = {
            'http_code': r'HTTP_CODE:\s*(\d+)',
            'dns': r'DNS:\s*([\d.]+)s',
            'tcp': r'TCP:\s*([\d.]+)s',
            'ssl': r'SSL:\s*([\d.]+)s',
            'ttfb': r'TTFB:\s*([\d.]+)s',
            'total': r'TOTAL:\s*([\d.]+)s'
        }
        
        for key, pattern in patterns.items():
            match = re.search(pattern, output)
            if match:
                data[key] = match.group(1)
        
        return data
        
    except subprocess.TimeoutExpired:
        return {
            'url': url,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'http_code': 'TIMEOUT',
            'dns': '-',
            'tcp': '-',
            'ssl': '-',
            'ttfb': '-',
            'total': '-'
        }
    except Exception:
        return {
            'url': url,
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'http_code': 'ERROR',
            'dns': '-',
            'tcp': '-',
            'ssl': '-',
            'ttfb': '-',
            'total': '-'
        }


def get_status(http_code):
    """根据HTTP状态码判断状态"""
    try:
        code = int(http_code)
        if code in [200, 301, 302, 304]:
            return '成功'
        elif code == 0:
            return '失败'
        else:
            return '成功' if 200 <= code < 400 else '失败'
    except (ValueError, TypeError):
        return '失败'


def save_batch_to_excel(all_rounds, output_file):
    """保存多轮结果到Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "拨测结果"
    
    # 表头
    headers = ['URL', '测试时间', 'HTTP状态码', 'DNS解析(s)', 'TCP连接(s)', 'SSL握手(s)', 'TTFB(s)', '总耗时(s)', '状态']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 写入数据
    row = 2
    for round_results in all_rounds:
        for result in round_results:
            ws.cell(row=row, column=1, value=result['url'])
            ws.cell(row=row, column=2, value=result['time'])
            ws.cell(row=row, column=3, value=result['http_code'])
            ws.cell(row=row, column=4, value=result['dns'])
            ws.cell(row=row, column=5, value=result['tcp'])
            ws.cell(row=row, column=6, value=result['ssl'])
            ws.cell(row=row, column=7, value=result['ttfb'])
            ws.cell(row=row, column=8, value=result['total'])
            ws.cell(row=row, column=9, value=get_status(result['http_code']))
            row += 1
    
    # 调整列宽
    column_widths = [40, 20, 12, 12, 12, 12, 12, 12, 8]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    
    wb.save(output_file)
    print(f"\n结果已保存到: {output_file}")


def run_test(round_num):
    """执行一轮测试，返回结果列表"""
    print(f"\n{'='*80}")
    print(f"第 {round_num} 轮拨测 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"共 {len(DOMAINS)} 个域名")
    print(f"{'='*80}")
    
    results = []
    for i, domain in enumerate(DOMAINS, 1):
        if not running:
            break
        print(f"[{i}/{len(DOMAINS)}] {domain}")
        # mymall 使用特定的 User-Agent
        if "mymall.hk.chinamobile.com" in domain:
            result = curl_test(domain, user_agent="hshhk/android/")
        else:
            result = curl_test(domain)
        results.append(result)
        print(f"    HTTP: {result['http_code']} | DNS: {result['dns']}s | TOTAL: {result['total']}s")
    
    return results


def main():
    global running
    
    # 注册信号处理
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    # 创建输出目录
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)
        print(f"创建输出目录: {OUTPUT_DIR}")
    
    print("="*80)
    print("域名拨测自动化脚本")
    print(f"执行间隔: 每 {INTERVAL_MINUTES} 分钟")
    print(f"汇总周期: 每 {BATCH_SIZE} 次拨测")
    print(f"域名数量: {len(DOMAINS)}")
    print(f"输出目录: {OUTPUT_DIR}")
    print("按 Ctrl+C 停止运行")
    print("="*80)
    
    round_num = 0
    all_rounds = []  # 存储多轮结果
    batch_start_time = None  # 记录批次开始时间
    
    while running:
        round_num += 1
        
        # 记录批次开始时间
        if batch_start_time is None:
            batch_start_time = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        results = run_test(round_num)
        if results:
            all_rounds.append(results)
        
        batch_round = len(all_rounds)
        print(f"\n当前批次进度: {batch_round}/{BATCH_SIZE}")
        
        # 达到批次大小，输出Excel
        if batch_round >= BATCH_SIZE:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_file = os.path.join(OUTPUT_DIR, f"curl_batch_{batch_start_time}_to_{timestamp}.xlsx")
            save_batch_to_excel(all_rounds, output_file)
            
            # 重置
            all_rounds = []
            batch_start_time = None
            print(f"\n{'='*80}")
            print(f"批次完成！开始新一轮批次...")
            print(f"{'='*80}")
        
        if not running:
            break
        
        # 等待下一轮
        print(f"\n下一轮将在 {INTERVAL_MINUTES} 分钟后执行...")
        print(f"当前时间: {datetime.now().strftime('%H:%M:%S')}")
        
        # 分段sleep，便于响应中断信号
        for _ in range(INTERVAL_MINUTES * 60):
            if not running:
                break
            time.sleep(1)
    
    # 退出前保存未完成的批次
    if all_rounds:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = os.path.join(OUTPUT_DIR, f"curl_batch_{batch_start_time}_to_{timestamp}_incomplete.xlsx")
        save_batch_to_excel(all_rounds, output_file)
        print(f"未完成批次已保存 ({len(all_rounds)}/{BATCH_SIZE} 轮)")
    
    print("\n脚本已停止运行")
    print(f"共完成 {round_num} 轮拨测")


if __name__ == "__main__":
    main()
