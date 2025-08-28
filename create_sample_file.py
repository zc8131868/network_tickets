#!/usr/bin/env python
"""
Script to create a sample.xlsx file entry in the DownloadFile model.
Run this script to set up the sample file for download.
"""

import os
import django
from pathlib import Path

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'network_tickets.settings')
django.setup()

from auto_tickets.models import DownloadFile

def create_sample_file_entry():
    """Create a DownloadFile entry for the sample.xlsx file"""
    
    # Check if sample file already exists
    if DownloadFile.objects.filter(title='Sample').exists():
        print("Sample file entry already exists in database.")
        existing = DownloadFile.objects.get(title='Sample')
        print(f"Existing entry: ID={existing.id}, File={existing.file}")
        return
    
    # Create a sample.xlsx file if it doesn't exist
    sample_file_path = Path('media/download_files/sample.xlsx')
    sample_file_path.parent.mkdir(parents=True, exist_ok=True)
    
    if not sample_file_path.exists():
        # Create a simple Excel file using openpyxl
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Content"
            
            # Add some sample data
            ws['A1'] = 'Sample Data'
            ws['B1'] = 'Source IP'
            ws['C1'] = 'Destination IP'
            ws['A2'] = 'Example'
            ws['B2'] = '192.168.1.1'
            ws['C2'] = '10.0.0.1'
            
            wb.save(sample_file_path)
            print(f"Created sample file at: {sample_file_path}")
        except ImportError:
            print("openpyxl not available. Please install it with: pip install openpyxl")
            return
    
    # Create DownloadFile entry
    try:
        download_file = DownloadFile.objects.create(
            title='Sample',
            file='download_files/sample.xlsx'
        )
        print(f"Created DownloadFile entry: ID={download_file.id}, Title={download_file.title}")
        print("Sample file is now available for download!")
    except Exception as e:
        print(f"Error creating DownloadFile entry: {e}")

if __name__ == "__main__":
    create_sample_file_entry()
